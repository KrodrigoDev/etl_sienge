from __future__ import annotations

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import date

# ─────────────────────────────────────────────────────────────────────────────
# PALETA DE CORES
# ─────────────────────────────────────────────────────────────────────────────

AZUL_ESCURO = "1F3864"
AZUL_MEDIO = "2E75B6"
AZUL_CLARO = "BDD7EE"
CINZA_LINHA = "F2F2F2"
VERDE = "375623"
VERDE_CLARO = "E2EFDA"
AMARELO = "FFE699"
LARANJA = "F4B942"
VERMELHO_CL = "FCE4D6"
BRANCO = "FFFFFF"
PRETO = "000000"

SCORE_CORES = {
    "exato": ("375623", "E2EFDA"),
    "alto": ("7F6000", "FFEB9C"),
    "medio": ("9C6500", "FCE4D6"),
    "baixo": ("843C0C", "FFC7CE"),
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────────────────────────────────────

def _borda(style="thin"):
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold=False, color=PRETO, size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _estilo_header(ws, row, col, valor, bg=AZUL_ESCURO, fg=BRANCO, size=10, bold=True, align="center"):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = _font(bold=bold, color=fg, size=size)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center")
    c.border = _borda()
    return c


def _estilo_dado(ws, row, col, valor, bg=BRANCO, bold=False, align="left", number_format=None):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = _font(bold=bold, color=PRETO)
    c.fill = _fill(bg)
    c.alignment = _align(align, "center", wrap=True)
    c.border = _borda()
    if number_format:
        c.number_format = number_format
    return c


def _mesclar_titulo(ws, row, col_ini, col_fim, valor, bg=AZUL_MEDIO, fg=BRANCO, size=12):
    ws.merge_cells(start_row=row, start_column=col_ini, end_row=row, end_column=col_fim)
    c = ws.cell(row=row, column=col_ini, value=valor)
    c.font = _font(bold=True, color=fg, size=size)
    c.fill = _fill(bg)
    c.alignment = _align("center", "center")
    c.border = _borda()
    return c


# ─────────────────────────────────────────────────────────────────────────────
# ABA: CAPA / RESUMO
# ─────────────────────────────────────────────────────────────────────────────

def _aba_resumo(wb, n_matched, n_only_giss, n_only_sienge, df_matched):
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.row_dimensions[1].height = 8

    # Cabeçalho institucional
    ws.merge_cells("B2:F2")
    c = ws.cell(row=2, column=2, value="CONCILIAÇÃO GISS × SIENGE")
    c.font = _font(bold=True, color=BRANCO, size=18)
    c.fill = _fill(AZUL_ESCURO)
    c.alignment = _align("center", "center")

    ws.merge_cells("B3:F3")
    c = ws.cell(row=3, column=2, value=f"Relatório gerado em: {date.today().strftime('%d/%m/%Y')}")
    c.font = _font(color=BRANCO, size=10)
    c.fill = _fill(AZUL_MEDIO)
    c.alignment = _align("center", "center")

    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 20

    # KPIs
    total = n_matched + n_only_giss + n_only_sienge
    kpis = [
        ("Conciliados", n_matched, VERDE_CLARO, VERDE),
        ("Apenas no GISS", n_only_giss, AMARELO, "7F6000"),
        ("Apenas no SIENGE", n_only_sienge, VERMELHO_CL, "843C0C"),
        ("Total processado", total, AZUL_CLARO, AZUL_ESCURO),
    ]

    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 22

    for i, (label, valor, bg, fg) in enumerate(kpis):
        col = 2 + i
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col)
        c = ws.cell(row=5, column=col, value=label)
        c.font = _font(bold=True, color=BRANCO, size=9)
        c.fill = _fill(AZUL_ESCURO)
        c.alignment = _align("center", "center")
        c.border = _borda()

        c2 = ws.cell(row=6, column=col, value=valor)
        c2.font = _font(bold=True, color=fg, size=22)
        c2.fill = _fill(bg)
        c2.alignment = _align("center", "center")
        c2.border = _borda()

        pct = valor / total if total else 0
        c3 = ws.cell(row=7, column=col, value=pct)
        c3.font = _font(color=fg, size=9)
        c3.fill = _fill(bg)
        c3.alignment = _align("center", "center")
        c3.border = _borda()
        c3.number_format = "0.0%"

    # Distribuição por score_label
    _mesclar_titulo(ws, 9, 2, 5, "Distribuição por Grau de Similaridade", AZUL_MEDIO)

    for j, label in enumerate(["Grau", "Qtd.", "% do total", "Descrição"]):
        _estilo_header(ws, 10, 2 + j, label, AZUL_ESCURO)

    descricoes = {
        "exato": "Todos os campos coincidem (score 90–100)",
        "alto": "Pequena divergência em valor ou data (80–89)",
        "medio": "Divergência moderada (70–79)",
        "baixo": "Poucos campos coincidem (< 70)",
    }

    if not df_matched.empty and "score_label" in df_matched.columns:
        dist = df_matched["score_label"].value_counts()
    else:
        dist = pd.Series(dtype=int)

    row = 11
    for label in ["exato", "alto", "medio", "baixo"]:
        qtd = dist.get(label, 0)
        bg = SCORE_CORES[label][1] if qtd > 0 else CINZA_LINHA
        _estilo_dado(ws, row, 2, label.capitalize(), bg, align="center")
        _estilo_dado(ws, row, 3, qtd, bg, align="center")
        pct_val = qtd / n_matched if n_matched else 0
        c = _estilo_dado(ws, row, 4, pct_val, bg, align="center")
        c.number_format = "0.0%"
        _estilo_dado(ws, row, 5, descricoes[label], bg)
        row += 1

    # Rodapé
    ws.row_dimensions[row + 1].height = 14
    ws.merge_cells(f"B{row + 2}:F{row + 2}")
    c = ws.cell(row=row + 2, column=2,
                value="Threshold mínimo: 90 pts  |  Tolerância de datas: 45 dias  |  Pesos: CNPJ 50 | Doc 30 | Valor 10 | Data 10")
    c.font = _font(color="595959", size=8)
    c.alignment = _align("center", "center")


# ─────────────────────────────────────────────────────────────────────────────
# ABA GENÉRICA DE DADOS
# ─────────────────────────────────────────────────────────────────────────────

COLUNAS_MATCHED = {
    "cnpj_empresa_giss": ("CNPJ Empresa", 20),
    "competencia_giss": ("Competência", 13),
    "cnpj_cpf_giss": ("CNPJ/CPF Prestador", 20),
    "prestador_giss": ("Prestador", 50),
    "nfs_giss": ("NFS-e GISS", 13),
    "valor_giss": ("Valor GISS (R$)", 16),
    "situacao_giss": ("Situação GISS", 14),
    "declaracao_giss": ("Declaração GISS", 16),
    "titulo_sienge": ("Título SIENGE", 14),
    "credor_sienge": ("Credor SIENGE", 60),
    "cnpj/cpf_sienge": ("CNPJ/CPF SIENGE", 20),
    "documento_sienge": ("Documento SIENGE", 18),
    "emissao_nf_sienge": ("Emissão NF SIENGE", 16),
    "valor_bruto_sienge": ("Valor Bruto SIENGE", 18),
    "score_similaridade": ("Score", 10),
    "score_label": ("Grau", 12),
}

COLUNAS_GISS = {
    "cnpj_empresa": ("CNPJ Empresa", 20),
    "empresa": ("Empresa", 70),
    "competencia": ("Competência", 13),
    "cnpj_cpf": ("CNPJ/CPF", 20),
    "prestador": ("Prestador", 70),
    "nfs": ("NFS-e", 13),
    "valor": ("Valor (R$)", 16),
    "situacao": ("Situação", 14),
    "declaracao": ("Declaração", 25),
    "emissao": ("Emissão", 20),
}

COLUNAS_SIENGE = {
    "titulo": ("Título", 14),
    "credor": ("Credor", 60),
    "cnpj/cpf": ("CNPJ/CPF", 20),
    "documento": ("Documento", 18),
    "emissao_nf": ("Emissão NF", 16),
    "valor_bruto": ("Valor Bruto (R$)", 18),
}


def _aba_dados(wb, nome_aba, df: pd.DataFrame, mapa_colunas: dict,
               titulo: str, cor_titulo=AZUL_MEDIO):
    ws = wb.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols_disponiveis = {k: v for k, v in mapa_colunas.items() if k in df.columns}
    colunas_orig = list(cols_disponiveis.keys())

    # Título
    n_cols = max(len(colunas_orig), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = _font(bold=True, color=BRANCO, size=13)
    c.fill = _fill(cor_titulo)
    c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 26

    # Cabeçalhos
    ws.row_dimensions[2].height = 22
    for j, col_orig in enumerate(colunas_orig, start=1):
        label, width = cols_disponiveis[col_orig]
        _estilo_header(ws, 2, j, label, AZUL_ESCURO)
        ws.column_dimensions[get_column_letter(j)].width = width

    # Dados
    for i, (_, row_data) in enumerate(df[colunas_orig].iterrows()):
        bg = CINZA_LINHA if i % 2 == 0 else BRANCO
        excel_row = i + 3

        # cor especial para score_label se existir
        score_bg = None
        if "score_label" in colunas_orig:
            lbl = row_data.get("score_label", "")
            if lbl in SCORE_CORES:
                score_bg = SCORE_CORES[lbl][1]

        for j, col_orig in enumerate(colunas_orig, start=1):
            val = row_data[col_orig]
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""

            cell_bg = bg
            # destaque nas colunas de score
            if col_orig == "score_label" and score_bg:
                cell_bg = score_bg
            elif col_orig == "score_similaridade" and score_bg:
                cell_bg = score_bg

            fmt = None
            align = "left"
            if col_orig in ("valor_giss", "valor_bruto_sienge", "valor", "valor_bruto"):
                try:
                    if isinstance(val, str):
                        val = (
                            val
                            .replace('.', '')
                            .replace(',', '.')
                            .strip()
                        )

                    val = float(val)

                except:
                    val = None

                # fmt = '#,##0.00'
            elif col_orig == "score_similaridade":
                align = "center"
            elif col_orig == "score_label":
                align = "center"
            elif "cnpj" in col_orig.lower():
                align = "center"
            elif "data" in col_orig.lower() or "emissao" in col_orig.lower() or "competencia" in col_orig.lower():
                align = "center"

            c = _estilo_dado(ws, excel_row, j, val, cell_bg, align=align, number_format=fmt)

        ws.row_dimensions[excel_row].height = 16

    # Filtro automático
    if len(colunas_orig) > 0 and len(df) > 0:
        ws.auto_filter.ref = (
            f"A2:{get_column_letter(len(colunas_orig))}{len(df) + 2}"
        )

    # Rodapé de contagem
    rodape_row = len(df) + 4
    ws.merge_cells(start_row=rodape_row, start_column=1, end_row=rodape_row, end_column=n_cols)
    c = ws.cell(row=rodape_row, column=1, value=f"Total de registros: {len(df)}")
    c.font = _font(bold=True, color=AZUL_ESCURO, size=9)
    c.fill = _fill(AZUL_CLARO)
    c.alignment = _align("right", "center")
    ws.row_dimensions[rodape_row].height = 18


# ─────────────────────────────────────────────────────────────────────────────
# FUNÇÃO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def gerar_relatorio_xlsx(
        df_matched: pd.DataFrame,
        df_only_giss: pd.DataFrame,
        df_only_sienge: pd.DataFrame,
        caminho_saida: str = "relatorio_conciliacao.xlsx",
):
    wb = Workbook()

    _aba_resumo(
        wb,
        n_matched=len(df_matched),
        n_only_giss=len(df_only_giss),
        n_only_sienge=len(df_only_sienge),
        df_matched=df_matched,
    )

    _aba_dados(
        wb,
        nome_aba="Conciliados",
        df=df_matched,
        mapa_colunas=COLUNAS_MATCHED,
        titulo="Registros Conciliados — GissOnline × Sienge",
        cor_titulo="375623",
    )

    _aba_dados(
        wb,
        nome_aba="Apenas GISS",
        df=df_only_giss,
        mapa_colunas=COLUNAS_GISS,
        titulo="Registros presentes somente no GissOnline Maceió",
        cor_titulo="7F6000",
    )

    _aba_dados(
        wb,
        nome_aba="Apenas SIENGE",
        df=df_only_sienge,
        mapa_colunas=COLUNAS_SIENGE,
        titulo="Registros presentes somente no Sienge",
        cor_titulo="843C0C",
    )

    wb.save(caminho_saida)
    print(f"Relatório salvo em: {caminho_saida}")

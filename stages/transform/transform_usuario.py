"""
stages/transform/transform_usuarios.py
-----------------------------------------------
Produz as seguintes tabelas para o modelo de usuários do SIENGE:

  ┌─ DIMENSÕES ──────────────────────────────────────────────────────────────┐
  │  dim_usuario              — atributos do usuário (cargo, tipo, flags)    │
  │  dim_acoes_sistema        — catálogo de ações disponíveis por sistema    │
  │  dim_perfil_usuario       — perfis de acesso atribuídos por usuário      │
  └──────────────────────────────────────────────────────────────────────────┘
  ┌─ FATOS / SATÉLITES (grain = usuário × entidade × função) ────────────────┐
  │  fato_acesso_usuario      — snapshot de engajamento (1 linha/usuário)    │
  │  dim_permissao            — permissões de ação no sistema (170k+ linhas) │
  │  fato_permissao_empresa   — autorização por empresa  × 6 funções         │
  │  fato_permissao_departamento — autorização por depto × 9 funções         │
  │  fato_permissao_obra      — autorização por obra     × 23 funções        │
  └──────────────────────────────────────────────────────────────────────────┘

Fontes
------
  cadastro_usuario_*.csv        — extraído via Selenium + BeautifulSoup
  relatorio_usuario.xlsx        — relatório SIENGE (cargo por usuário)
  permissao_usuario*.csv        — matriz de permissões por usuário/ação
  permissoes_sistema*.xlsx      — catálogo de ações disponíveis por sistema
  perfil_usuario*.xlsx          — perfis atribuídos por usuário
  permissao_empresa*.xlsx       — autorizações por empresa (6 funções)
  permissao_departamento*.xlsx  — autorizações por departamento (9 funções)
  permissao_obra*.xlsx          — autorizações por obra (23 funções)

Relacionamentos Power BI
------------------------
  dim_usuario[id_usuario]  ──1:1──  fato_acesso_usuario[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  dim_permissao[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  dim_perfil_usuario[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  fato_permissao_empresa[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  fato_permissao_departamento[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  fato_permissao_obra[id_usuario]
  dim_acoes_sistema[codigo] ─M:1──  dim_permissao[acao_id]
  dim_mapeamento_cargo_perfil[cargo]           → dim_usuario[cargo]
  dim_mapeamento_cargo_perfil[perfil_esperado] → dim_perfil_usuario[perfil_codigo]
  dim_aderencia_perfil[id_usuario]             → dim_usuario[id_usuario]

  Hub central: dim_usuario. Todos os satélites se ligam via id_usuario.
  fato_acesso_usuario é um satélite de engajamento com grain 1:1.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from stages.transform.utils.normalizer import salvar_tabela

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

pasta_origem = Path(__file__).resolve().parents[2]

INPUT_DIR = pasta_origem / "stages" / "transform" / "input"
OUTPUT_DIR = pasta_origem / "stages" / "transform" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

DOMINIOS_INTERNOS = {"telesil.com.br", "telesilengenharia.com.br"}

CODIGOS_SISTEMA = {
    "ADMIN", "SUPER", "TESTE", "TESTE1", "PERMISSAO",
    "RPATELESIL", "TITELESIL",
    "LCT01", "LCT02", "LCT03", "LCT04",
    "JAPRENDIZ", "NG7", "NOTASMARKETING",
}

# ── Mapas de funções por escopo ───────────────────────────────────────────────

FUNCOES_EMPRESA_MAP: dict[int, str] = {
    1: "Acesso aos dados da empresa",
    2: "Cadastrar títulos a pagar autorizados automaticamente",
    3: "Cadastrar contas a pagar sem limite mínimo de dias",
    4: "Cadastrar contratos de incorporação",
    5: "Cadastrar baixa do contas a receber de recebimento",
    6: "Cadastrar contas a pagar com vencimento fora do período",
}

FUNCOES_DEPARTAMENTO_MAP: dict[int, str] = {
    1: "Acesso - Gerencial Financeiro",
    2: "Acesso - Autorização Suprimentos",
    3: "Manutenção - Orçamento Empresarial",
    4: "Manutenção - Financeiro",
    5: "Acesso - Pagamento Escritural",
    6: "Inclusão - Previsão Financeira Pedidos de Compra",
    7: "Inclusão - Previsão Financeira Contratos",
    8: "Inclusão - Notas Fiscais de Compra",
    9: "Inclusão - Liberações de Medições",
}

FUNCOES_OBRA_MAP: dict[int, str] = {
    1: "Acesso - Orçamento",
    2: "Acesso - Planejamento",
    3: "Acesso - Acompanhamento",
    4: "Acesso - Controle de Mão de Obra",
    5: "Acesso - Gerencial de Obras",
    6: "Acesso - Gerencial Financeiro",
    7: "Acesso - Gerencial Suprimentos",
    8: "Manutenção - Compras",
    9: "Manutenção - Contratos",
    10: "Manutenção - Medições",
    11: "Manutenção - Estoque",
    12: "Manutenção - Financeiro",
    13: "Manutenção - Gestão da Qualidade",
    14: "Manutenção - Administrativo",
    15: "Manutenção - Comercial",
    16: "Manutenção - Orçamento Empresarial",
    17: "Manutenção - Diário de Obra",
    18: "Manutenção - Integração SAP",
    19: "Acesso - Custo Orçado e Incorrido",
    20: "Acesso - Locações de Equipamentos",
    21: "Acesso - Nota Fiscal Eletrônica",
    22: "Exclusão de Anexos",
    23: "Acesso - Apoio",
}

# ── Mapeamento cargo → perfis esperados ──────────────────────────────────────
# Opção 2: tabela explícita. Atualizar quando novos cargos ou perfis surgirem.
# Um cargo pode ter múltiplos perfis válidos (acumulação de funções é legítima).
# Os códigos de perfil devem bater exatamente com dim_perfil_usuario[perfil_codigo].
MAPEAMENTO_CARGO_PERFIL: dict[str, list[str]] = {
    # ── Engenharia / Obras ────────────────────────────────────────────────────
    "Engenheiro de obras": ["ENG OBRAS", "GESTORES OBRAS"],
    "Analista de Obra": ["GESTORES OBRAS", "ANALIST ENGENHARIA"],
    "Auxiliar de Engenharia": ["ANALIST ENGENHARIA", "ENGENHARIA"],
    "Coordenador de Engenharia": ["GESTORES OBRAS", "ANALIST ENGENHARIA"],
    "Gerente de engenharia": ["GESTORES OBRAS"],
    "Diretor técnico": ["GESTORES OBRAS"],
    "Diretor de operações": ["GESTORES OBRAS"],
    "Técnico de Edificações": ["ANALIST ENGENHARIA", "ENGENHARIA"],
    "Engenheiro Eletricista": ["ANALIST ENGENHARIA", "ENGENHARIA"],
    "Administrativo de obra": ["GESTORES OBRAS"],

    # ── Planejamento ──────────────────────────────────────────────────────────
    "Analista de Planejamento": ["PLANEJAMENTO OBRAS"],
    "Gerente de planejamento": ["PLANEJAMENTO OBRAS", "GESTORES OBRAS"],
    "Gerente de projetos": ["PLANEJAMENTO OBRAS", "GESTORES OBRAS"],

    # ── Suprimentos / Compras ─────────────────────────────────────────────────
    "Almoxarife": ["ALMOXARIFE"],
    "Auxiliar de Almoxarife": ["ALMOXARIFE"],
    "Comprador": ["SUPRIMENTOS", "SUPRIMENTOS NF"],
    "Coordenador de Suprimentos": ["SUPRIMENTOS", "SUPRIMENTOS NF"],
    "Gerente de suprimentos": ["SUPRIMENTOS", "SUPRIMENTOS NF"],
    "Gerente de contratos": ["SUPRIMENTOS"],
    "Analista de Logistica": ["SUPRIMENTOS"],

    # ── Financeiro / Contabilidade ────────────────────────────────────────────
    "Analista financeiro": ["FIN CONT A PAG", "FINANCEIRO  ADM"],
    "Contador": ["FIN CONT A PAG", "FINANCEIRO  ADM"],
    "Coordenador financeiro": ["FIN CONT A PAG", "FINANCEIRO  ADM"],
    "Diretor financeiro": ["FIN CONT A PAG", "FINANCEIRO  ADM"],
    "Controller (eventualmente)": ["FIN CONT A PAG", "FINANCEIRO  ADM"],

    # ── Administrativo / RH ───────────────────────────────────────────────────
    "Assistente administrativo": ["FINANCEIRO  ADM"],
    "Analista Administrativo (Trainee)": ["FINANCEIRO  ADM"],
    "Analista DP": ["FINANCEIRO  ADM"],
    "Assistente DP": ["FINANCEIRO  ADM"],
    "Gerente de DP": ["FINANCEIRO  ADM"],
    "Diretor administrativo": ["FINANCEIRO  ADM"],

    # ── Comercial / Marketing / Incorporação ──────────────────────────────────
    "Analista comercial": ["COMERCIAL"],
    "Coordenadora de Repasse": ["COMERCIAL"],
    "Relacionamento com o Cliente": ["COMERCIAL"],
    "Assistente de Marketing": ["COMERCIAL"],
    "Diretoria de Marketing": ["COMERCIAL"],
    "Gerente de incorporação": ["COMERCIAL", "VERTICAL"],
    "Orçamentista": ["VERTICAL"],

    # ── Dados / TI ────────────────────────────────────────────────────────────
    "Analista de Dados": [],  # sem perfil padrão definido — avaliar
    "TI": [],  # sem perfil padrão definido — avaliar

    # ── Sem classificação clara ───────────────────────────────────────────────
    "Outro": [],  # heterogêneo — não mapear
}

# ── Normalização: nome do perfil no CSV → perfil_codigo na dim_perfil_usuario ─
# O relatório permissao_perfil usa nomes longos; a dim usa códigos curtos.
# Atualizar se novos perfis forem criados no SIENGE.
NORMALIZACAO_PERFIL_CSV: dict[str, str] = {
    "ALMOXARIFE": "ALMOXARIFE",
    "ANALISTA DE ENGENHARIA": "ANALIST ENGENHARIA",
    "COMERCIAL": "COMERCIAL",
    "ENGENHARIA": "ENGENHARIA",
    "ENGENHEIRO DE OBRAS": "ENG OBRAS",
    "FINANCEIRO CONTAS A PAGAR": "FIN CONT A PAG",
    "FINANCEIRO ESCRITORIO CENTRAL": "FINANCEIRO  ADM",
    "FINANCEIRO - OBRAS": "FINANCEIRO  ADM",
    "FORTEMIX": "FORTEMIX",  # sem equivalente na dim — novo perfil
    "GESTORES OBRAS": "GESTORES OBRAS",
    "PERFIL VERTICAL": "VERTICAL",
    "PLANEJAMENTO OBRAS": "PLANEJAMENTO OBRAS",
    "SUPRIMENTOS": "SUPRIMENTOS",
    "SUPRIMENTOS - NF": "SUPRIMENTOS NF",
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, format="%d/%m/%Y", errors="coerce")


def _dominio_email(email: str | float) -> str:
    if not isinstance(email, str) or "@" not in email:
        return "sem_email"
    return email.strip().lower().split("@")[-1]


def _tipo_usuario(row: pd.Series) -> str:
    codigo = str(row.get("codigo", "")).upper()
    dominio = row.get("dominio_email", "")
    if codigo in CODIGOS_SISTEMA:
        return "sistema_teste"
    if dominio in DOMINIOS_INTERNOS:
        return "interno"
    if dominio == "sem_email":
        return "sem_email"
    return "externo"


def _faixa_inatividade(dias: float | None) -> str:
    if dias is None or np.isnan(dias):
        return "Nunca acessou"
    if dias <= 7:   return "Ativo (≤7d)"
    if dias <= 30:  return "Recente (8-30d)"
    if dias <= 60:  return "Alerta (31-60d)"
    if dias <= 90:  return "Crítico (61-90d)"
    return "Inativo (>90d)"


def _faixa_antiguidade(dias: float | None) -> str:
    if dias is None or np.isnan(dias):
        return "Desconhecido"
    if dias <= 30:   return "Novo (≤30d)"
    if dias <= 180:  return "Recente (31-180d)"
    if dias <= 365:  return "Intermediário (181-365d)"
    if dias <= 730:  return "Experiente (1-2 anos)"
    return "Veterano (>2 anos)"


def _extrair_acao_id(acao: str) -> str | None:
    if not isinstance(acao, str):
        return None
    match = re.search(r"\((\d+)\)$", acao)
    return match.group(1) if match else None


def _parser_blocos_usuario(
        rows: list[tuple],
        funcoes_map: dict[int, str],
        entidade_col_nome: str,
        sep_entidade: str = "-",
) -> list[dict]:
    """
    Parser genérico para relatórios SIENGE com estrutura de blocos por usuário.

    Todos os relatórios de autorização (empresa, departamento, obra) seguem
    exatamente o mesmo layout:

      row+0: 'Usuário' | (col 1 vazia) | CODIGO_USUARIO   ← col 2, sem separador
      row+1: 'Funções' | legenda textual (ignorada)
      row+2: '<label>' | 'Funções'
      row+3:  None...  | 1 | 2 | 3 ...   ← funcao_id por col_index
      row+4+: ENTIDADE_RAW | None/Sim... por coluna

    O código do usuário está sempre na col 2 (sem separador ' - ').
    O separador da entidade varia por arquivo:
      empresa:      ' - '  →  '1 - TELESIL ENGENHARIA LTDA'
      departamento: '-'    →  '4-DEPARTAMENTO DE ENGENHARIA'
      obra:         '-'    →  '2125-EDIFICIO DOM ANTONIO'

    Parâmetros
    ----------
    rows             : linhas do worksheet já lidas
    funcoes_map      : dicionário funcao_id → descrição
    entidade_col_nome: prefixo da chave de código ('empresa_codigo', etc.)
    sep_entidade     : separador código/nome da entidade
    """
    usuario_starts = [i for i, r in enumerate(rows) if r[0] == "Usuário"]
    registros: list[dict] = []
    nome_col = entidade_col_nome.replace("_codigo", "_nome")

    for bloco_idx, usr_row_i in enumerate(usuario_starts):
        # Código do usuário: col 2, sem separador
        codigo_usuario = (
            str(rows[usr_row_i][2]).strip().upper()
            if rows[usr_row_i][2] else ""
        )

        codigo_usuario = codigo_usuario.split('-', maxsplit=1)[0].strip()

        # col_index → funcao_id (linha usr_row_i + 3)
        num_row_i = usr_row_i + 3
        num_row = rows[num_row_i]
        col_to_funcao = {
            col_i: int(v)
            for col_i, v in enumerate(num_row)
            if isinstance(v, (int, float)) and int(v) in funcoes_map
        }

        fim = (
            usuario_starts[bloco_idx + 1]
            if bloco_idx + 1 < len(usuario_starts)
            else len(rows)
        )

        for row in rows[num_row_i + 1: fim]:
            entidade_raw = row[0]
            if not isinstance(entidade_raw, str) or not entidade_raw.strip():
                continue
            # Ignorar cabeçalhos estruturais do relatório
            if entidade_raw.strip() in ("Obra", "Empresa", "Departamento", "Funções", "Usuário"):
                continue
            # Ignorar timestamps que o SIENGE injeta no final do arquivo
            if re.match(r"^\d{2}/\d{2}/\d{4}", str(entidade_raw).strip()):
                continue

            partes = str(entidade_raw).split(sep_entidade, 1)
            entidade_codigo = partes[0].strip()
            entidade_nome = partes[1].strip() if len(partes) > 1 else entidade_raw.strip()

            for col_i, funcao_id in col_to_funcao.items():
                val = row[col_i] if col_i < len(row) else None
                tem_acesso = str(val).strip().lower() == "sim"
                registros.append({
                    "codigo_usuario": codigo_usuario,
                    entidade_col_nome: entidade_codigo,
                    nome_col: entidade_nome,
                    "funcao_id": funcao_id,
                    "funcao_nome": funcoes_map[funcao_id],
                    "tem_acesso": tem_acesso,
                })

    return registros


def _enriquecer_com_id(
        df: pd.DataFrame,
        codigo_para_id: pd.DataFrame,
        nome_tabela: str,
) -> pd.DataFrame:
    """
    Faz o join de qualquer satélite de autorização com o lookup
    codigo_usuario → id_usuario e posiciona id_usuario na primeira coluna.
    Reporta códigos sem match para diagnóstico.
    """
    df = df.merge(codigo_para_id, on="codigo_usuario", how="left")
    sem_match_mask = df["id_usuario"].isna()
    if sem_match_mask.any():
        orphans = df.loc[sem_match_mask, "codigo_usuario"].unique()
        print(f"  {nome_tabela}: {len(orphans)} código(s) sem match: "
              f"{list(orphans)[:5]}")
    df.insert(0, "id_usuario", df.pop("id_usuario"))
    return df


# ─────────────────────────────────────────────────────────────────────────────
# LEITURA DAS FONTES
# ─────────────────────────────────────────────────────────────────────────────

def _ler_cadastro(input_dir: Path) -> pd.DataFrame:
    arquivos = list((input_dir / "usuario").glob("cadastro_usuario_*.csv"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum cadastro_usuario_*.csv encontrado em {input_dir / 'usuario'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo cadastro: {arquivo.name}")
    df = pd.read_csv(arquivo, encoding="utf-8-sig")
    df = df.drop_duplicates(subset="codigo", keep="first").reset_index(drop=True)
    return df


def _ler_relatorio(input_dir: Path) -> pd.DataFrame:
    arquivo = input_dir / "usuario" / "relatorio_usuario.xlsx"
    if not arquivo.exists():
        raise FileNotFoundError(f"Relatório não encontrado: {arquivo}")
    print(f"  Lendo relatório: {arquivo.name}")
    df = pd.read_excel(arquivo, header=4)
    df = df[["Usuário", "Nome", "Email", "Cargo", "Admin",
             "Data de ativação", "Data de desativação"]].copy()
    df.columns = [
        "codigo", "nome_relatorio", "email_relatorio", "cargo",
        "admin_relatorio", "data_ativacao_rel", "data_desativacao_rel",
    ]
    df = df[df["codigo"].notna()].reset_index(drop=True)
    df["codigo"] = df["codigo"].astype(str).str.strip().str.upper()
    df["admin_relatorio"] = df["admin_relatorio"].fillna("").astype(str).str.strip()
    return df


def _ler_acoes_sistema(input_dir: Path) -> pd.DataFrame:
    arquivos = list((input_dir / "reference").glob("permissoes_sistema*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum permissoes_sistema*.xlsx encontrado em {input_dir / 'reference'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo catálogo de ações: {arquivo.name}")
    df = pd.read_excel(arquivo, header=None)
    col0 = df.iloc[:, 0]
    mask = col0.astype(str).str.strip().eq("Sistema")
    df["sistema_temp"] = None
    for idx in df.index[mask]:
        if idx + 1 in df.index:
            df.loc[idx + 1:, "sistema_temp"] = df.iloc[idx + 1, 0]
    df["sistema"] = df["sistema_temp"].ffill()
    df = df[~col0.isin(["Sistema", "Código"])]
    df = df.drop(columns=["sistema_temp", 1, 3]).reset_index(drop=True)
    df.columns = ["codigo", "acao", "sistema"]
    df.dropna(subset=["acao"], inplace=True)

    df['cod_acao_pesquisa'] = df['codigo'].apply(lambda x: f'Nº {x}')

    return df


def _ler_permissao_usuario(input_dir: Path) -> pd.DataFrame:
    arquivos = list((input_dir / "usuario").glob("permissao_usuario*.csv"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum permissao_usuario*.csv encontrado em {input_dir / 'usuario'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo permissões de ação: {arquivo.name}")
    df = pd.read_csv(arquivo, sep=";")
    df = df.rename(columns={"Unnamed: 0": "acao"})
    df = df.drop(index=df.loc[df["acao"] == "Todas as ações"].index).reset_index(drop=True)
    df_melt = df.melt(id_vars="acao", var_name="usuario", value_name="tem_permissao")
    df_melt["usuario"] = df_melt["usuario"].astype(str).str.strip().str.upper()
    df_melt["tem_permissao"] = (
        df_melt["tem_permissao"].astype(str).str.strip().str.lower()
        .isin(["true", "1", "sim", "x"])
    )
    df_melt["acao_id"] = df_melt["acao"].apply(_extrair_acao_id)
    df_melt.dropna(subset=["acao_id"], inplace=True)
    return df_melt


def _ler_perfil_usuario(input_dir: Path) -> pd.DataFrame:
    """
    Lê o XLSX de perfis de acesso por usuário.

    Estrutura (diferente dos demais — sem matriz Sim/Não):
      row+0: 'Usuário' | (col 1 vazia) | CODIGO_USUARIO   ← só código, sem nome
      row+1: (vazio)
      row+2: 'Código'  | (col 1 vazia) | 'Nome'           ← cabeçalho
      row+3..N: CODIGO_PERFIL | (col 1) | NOME_PERFIL

    Grain: usuário × perfil (lista simples, sem funções ou Sim/Não).
    """
    arquivos = list((input_dir / "usuario").glob("perfil_usuario*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum perfil_usuario*.xlsx encontrado em {input_dir / 'usuario'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo perfis de usuário: {arquivo.name}")

    wb = load_workbook(arquivo, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    usuario_starts = [i for i, r in enumerate(rows) if r[0] == "Usuário"]
    registros: list[dict] = []

    for bloco_idx, usr_row_i in enumerate(usuario_starts):
        codigo_usuario = (
            str(rows[usr_row_i][2]).strip().upper()
            if rows[usr_row_i][2] else ""
        )
        fim = (
            usuario_starts[bloco_idx + 1]
            if bloco_idx + 1 < len(usuario_starts)
            else len(rows)
        )
        for row in rows[usr_row_i + 1: fim]:
            col0_val = row[0]
            col2_val = row[2]
            # Pular cabeçalhos e linhas vazias
            if not col0_val or str(col0_val).strip() in ("Código", "Usuário", ""):
                continue
            registros.append({
                "codigo_usuario": codigo_usuario,
                "perfil_codigo": str(col0_val).strip().upper(),
                "perfil_nome": str(col2_val).strip() if col2_val else str(col0_val).strip(),
            })

    print(f"    {len(usuario_starts)} usuários | {len(registros)} vínculos usuário-perfil")
    return pd.DataFrame(registros)


def _ler_permissao_perfil(input_dir: Path) -> pd.DataFrame:
    """
    Lê o CSV de permissões por perfil (relatório SIENGE).

    Estrutura idêntica à permissao_usuario:
      col 0  : nome da ação (ex: "FIN-CPG-Consultar (9102)")
      col 1+N: perfil (nome longo), valor "Sim"/"Não"

    Diferenças em relação a permissao_usuario:
      - Colunas são perfis (não usuários)
      - Valores são "Sim"/"Não" (não True/False)
      - Nomes das colunas usam nomes longos que diferem dos códigos da dim

    Normalização:
      Os nomes de colunas são mapeados via NORMALIZACAO_PERFIL_CSV para os
      perfil_codigo usados em dim_perfil_usuario, permitindo o join direto.

    Retorna dim_permissao_perfil (grain = perfil_codigo × acao_id):
        perfil_csv | perfil_codigo | acao_id | tem_permissao
    """
    arquivos = list((input_dir / "usuario").glob("permissao_perfil*.csv"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum permissao_perfil*.csv encontrado em {input_dir / 'reference'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo permissões por perfil: {arquivo.name}")

    df = pd.read_csv(arquivo, sep=";", encoding="utf-8-sig")
    df = df.rename(columns={"Unnamed: 0": "acao"})

    # Remover linha "Todas as ações" e linhas sem acao_id válido
    df = df.drop(index=df.loc[df["acao"] == "Todas as ações"].index).reset_index(drop=True)

    # Melt: uma linha por (acao, perfil)
    df_melt = df.melt(id_vars="acao", var_name="perfil_csv", value_name="tem_permissao")

    # Normalizar nome do perfil: strip + upper
    df_melt["perfil_csv"] = df_melt["perfil_csv"].astype(str).str.strip().str.upper()

    # "Sim"/"Não" → booleano
    df_melt["tem_permissao"] = (
            df_melt["tem_permissao"].astype(str).str.strip().str.lower() == "sim"
    )

    # Extrair acao_id do nome da ação
    df_melt["acao_id"] = df_melt["acao"].apply(_extrair_acao_id)
    df_melt.dropna(subset=["acao_id"], inplace=True)

    # Mapear nome longo → perfil_codigo (chave de join com dim_perfil_usuario)
    df_melt["perfil_codigo"] = df_melt["perfil_csv"].map(NORMALIZACAO_PERFIL_CSV)

    sem_mapa = df_melt["perfil_codigo"].isna().sum()
    if sem_mapa:
        novos = df_melt[df_melt["perfil_codigo"].isna()]["perfil_csv"].unique()
        print(f" {len(novos)} perfil(is) sem mapeamento em NORMALIZACAO_PERFIL_CSV: "
              f"{list(novos)} — adicionar ao dicionário")

    print(f"    {df_melt['perfil_csv'].nunique()} perfis | "
          f"{df_melt['acao_id'].nunique()} ações | "
          f"{df_melt['tem_permissao'].sum():,} acessos concedidos")

    return df_melt[["perfil_csv", "perfil_codigo", "acao_id", "tem_permissao"]]


def _ler_permissao_empresa(input_dir: Path) -> pd.DataFrame:
    """
    Lê autorizações por empresa. Separador da entidade: ' - '
    Ex: '1 - TELESIL ENGENHARIA LTDA'  →  codigo='1', nome='TELESIL ENGENHARIA LTDA'
    Funções: 6 (FUNCOES_EMPRESA_MAP)
    """
    arquivos = list((input_dir / "usuario").glob("permissao_empresa*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum permissao_empresa*.xlsx encontrado em {input_dir / 'usuario'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo permissões por empresa: {arquivo.name}")
    wb = load_workbook(arquivo, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    registros = _parser_blocos_usuario(
        rows, FUNCOES_EMPRESA_MAP, "empresa_codigo", sep_entidade=" - "
    )
    n_usr = len([i for i, r in enumerate(rows) if r[0] == "Usuário"])
    print(f"    {n_usr} usuários | {len(registros)} registros")
    return pd.DataFrame(registros)


def _ler_permissao_departamento(input_dir: Path) -> pd.DataFrame:
    """
    Lê autorizações por departamento. Separador da entidade: '-'
    Ex: '4-DEPARTAMENTO DE ENGENHARIA - OBRAS PRIVADAS'
        → codigo='4', nome='DEPARTAMENTO DE ENGENHARIA - OBRAS PRIVADAS'
    Funções: 9 (FUNCOES_DEPARTAMENTO_MAP)
    """
    arquivos = list((input_dir / "usuario").glob("permissao_departamento*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum permissao_departamento*.xlsx encontrado em {input_dir / 'usuario'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo permissões por departamento: {arquivo.name}")
    wb = load_workbook(arquivo, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    registros = _parser_blocos_usuario(
        rows, FUNCOES_DEPARTAMENTO_MAP, "departamento_codigo", sep_entidade="-"
    )
    n_usr = len([i for i, r in enumerate(rows) if r[0] == "Usuário"])
    print(f"    {n_usr} usuários | {len(registros)} registros")
    return pd.DataFrame(registros)


def _ler_permissao_obra(input_dir: Path) -> pd.DataFrame:
    """
    Lê autorizações por obra. Separador da entidade: '-'
    Ex: '2125-EDIFICIO DOM ANTONIO - VENDAS/MARKETING'
        → codigo='2125', nome='EDIFICIO DOM ANTONIO - VENDAS/MARKETING'
    Funções: 23 (FUNCOES_OBRA_MAP)
    """
    arquivos = list((input_dir / "usuario").glob("permissao_obra*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum permissao_obra*.xlsx encontrado em {input_dir / 'usuario'}"
        )
    arquivo = max(arquivos, key=lambda p: p.stat().st_mtime)
    print(f"  Lendo permissões por obra: {arquivo.name}")
    wb = load_workbook(arquivo, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    registros = _parser_blocos_usuario(
        rows, FUNCOES_OBRA_MAP, "obra_codigo", sep_entidade="-"
    )
    n_usr = len([i for i, r in enumerate(rows) if r[0] == "Usuário"])
    print(f"    {n_usr} usuários | {len(registros)} registros")
    return pd.DataFrame(registros)


def _gerar_mapeamento_cargo_perfil() -> pd.DataFrame:
    """
    Converte MAPEAMENTO_CARGO_PERFIL em dim_mapeamento_cargo_perfil.

    Grain: 1 linha por (cargo, perfil_esperado).
    Cargos com lista vazia geram 1 linha com perfil_esperado=None
    para permitir identificar gaps no Power BI.

    Relacionamentos sugeridos no Power BI:
      dim_mapeamento_cargo_perfil[cargo]           → dim_usuario[cargo]
      dim_mapeamento_cargo_perfil[perfil_esperado] → dim_perfil_usuario[perfil_codigo]
    """
    rows = []
    for cargo, perfis in MAPEAMENTO_CARGO_PERFIL.items():
        if perfis:
            for perfil in perfis:
                rows.append({
                    "cargo": cargo,
                    "perfil_esperado": perfil,
                    "mapeamento_definido": True,
                })
        else:
            rows.append({
                "cargo": cargo,
                "perfil_esperado": None,
                "mapeamento_definido": False,
            })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# PONTO DE ENTRADA
# ─────────────────────────────────────────────────────────────────────────────

def executar(input_dir: Path = INPUT_DIR, output_dir: Path = OUTPUT_DIR) -> None:
    hoje = pd.Timestamp(date.today())

    # ── 1. Leitura ────────────────────────────────────────────────────────────
    print("\n── 1. Leitura ──────────────────────────────────────────────────────")
    df_cad = _ler_cadastro(input_dir)
    df_rel = _ler_relatorio(input_dir)
    df_perm = _ler_permissao_usuario(input_dir)
    dim_acoes = _ler_acoes_sistema(input_dir)
    perfil_raw = _ler_perfil_usuario(input_dir)
    perm_perfil_raw = _ler_permissao_perfil(input_dir)
    empresa_raw = _ler_permissao_empresa(input_dir)
    depto_raw = _ler_permissao_departamento(input_dir)
    obra_raw = _ler_permissao_obra(input_dir)

    print(f"\n  Cadastro:            {len(df_cad):,} registros")
    print(f"  Relatório:           {len(df_rel):,} registros")
    print(f"  Permissão (ações):   {len(df_perm):,} registros")
    print(f"  Perfis:              {len(perfil_raw):,} registros")
    print(f"  Perm. empresa:       {len(empresa_raw):,} registros")
    print(f"  Perm. departamento:  {len(depto_raw):,} registros")
    print(f"  Perm. obra:          {len(obra_raw):,} registros")
    print(f"  Perm. perfil:        {len(perm_perfil_raw):,} registros")

    # ── 2. Join cadastro + relatório ──────────────────────────────────────────
    print("\n── 2. Join cadastro + relatório ────────────────────────────────────")
    df_cad["codigo"] = df_cad["codigo"].astype(str).str.strip().str.upper()
    df = df_cad.merge(
        df_rel[["codigo", "cargo", "admin_relatorio"]],
        on="codigo", how="left",
    )
    apenas_rel = set(df_rel["codigo"]) - set(df_cad["codigo"])
    if apenas_rel:
        print(f"  {len(apenas_rel)} usuários só no relatório (ignorados): "
              f"{list(apenas_rel)[:5]}")
    print(f"  Com cargo preenchido: {df['cargo'].notna().sum():,} / {len(df):,}")

    # ── 3. Conversão de tipos ─────────────────────────────────────────────────
    print("\n── 3. Conversão de tipos ───────────────────────────────────────────")
    df["data_ativacao"] = _parse_date(df["data_ativacao"])
    df["data_desativacao"] = _parse_date(df["data_desativacao"])
    df["data_ultimo_acesso"] = _parse_date(df["data_ultimo_acesso"])
    df["administrador"] = df["administrador"].astype(str).str.strip().str.lower() == "true"

    # ── 4. Campos derivados ───────────────────────────────────────────────────
    print("\n── 4. Campos derivados ─────────────────────────────────────────────")
    df["dominio_email"] = df["email"].apply(_dominio_email)
    df["tipo_usuario"] = df.apply(_tipo_usuario, axis=1)
    df["dias_sem_acesso"] = (hoje - df["data_ultimo_acesso"]).dt.days.astype("float64")
    df["antiguidade_dias"] = (hoje - df["data_ativacao"]).dt.days.astype("float64")

    df["flag_ativo"] = df["data_desativacao"].isna()
    df["flag_nunca_acessou"] = df["data_ultimo_acesso"].isna()
    df["flag_inativo_30d"] = df["dias_sem_acesso"] > 30
    df["flag_inativo_60d"] = df["dias_sem_acesso"] > 60
    df["flag_inativo_90d"] = df["dias_sem_acesso"] > 90
    df["flag_acessou_hoje"] = df["data_ultimo_acesso"].dt.normalize() == hoje
    df["flag_sistema_teste"] = df["tipo_usuario"] == "sistema_teste"
    df["flag_externo"] = df["tipo_usuario"] == "externo"
    df["flag_admin"] = df["administrador"]
    df["flag_nunca_acessou_ativo"] = df["flag_nunca_acessou"] & df["flag_ativo"]

    df["faixa_inatividade"] = df["dias_sem_acesso"].apply(_faixa_inatividade)
    df["faixa_antiguidade"] = df["antiguidade_dias"].apply(_faixa_antiguidade)

    df["status_engajamento"] = np.select(
        [
            df["flag_nunca_acessou"] & df["flag_ativo"],
            df["dias_sem_acesso"] <= 7,
            df["dias_sem_acesso"] <= 30,
            df["dias_sem_acesso"] <= 90,
            df["dias_sem_acesso"] > 90,
            ~df["flag_ativo"],
        ],
        ["Nunca acessou", "Ativo recente", "Ativo mensal",
         "Em alerta", "Inativo", "Desativado"],
        default="Desconhecido",
    )

    print(f"  Usuários ativos:          {df['flag_ativo'].sum():,}")
    print(f"  Nunca acessaram (ativos): {df['flag_nunca_acessou_ativo'].sum():,}")
    print(f"  Inativos >30d:            {df['flag_inativo_30d'].sum():,}")
    print(f"  Inativos >90d:            {df['flag_inativo_90d'].sum():,}")
    print(f"  Externos:                 {df['flag_externo'].sum():,}")
    print(f"  Administradores:          {df['flag_admin'].sum():,}")

    # ── 5. dim_usuario ────────────────────────────────────────────────────────
    print("\n── 5. dim_usuario ──────────────────────────────────────────────────")
    dim_usuario = df[[
        "codigo", "nome", "email", "cargo", "administrador",
        "provedor_identidade", "dominio_email", "tipo_usuario",
        "data_ativacao", "data_desativacao",
        "flag_ativo", "flag_admin", "flag_externo", "flag_sistema_teste",
        "faixa_antiguidade", "antiguidade_dias",
    ]].copy()
    dim_usuario.insert(0, "id_usuario", range(1, len(dim_usuario) + 1))
    print(f"  dim_usuario: {dim_usuario.shape}")

    # ── 6. fato_acesso_usuario ────────────────────────────────────────────────
    print("\n── 6. fato_acesso_usuario ──────────────────────────────────────────")
    fato = df[[
        "codigo", "nome", "cargo", "tipo_usuario", "dominio_email",
        "provedor_identidade", "data_ativacao", "data_desativacao",
        "data_ultimo_acesso", "dias_sem_acesso", "antiguidade_dias",
        "flag_ativo", "flag_nunca_acessou", "flag_nunca_acessou_ativo",
        "flag_inativo_30d", "flag_inativo_60d", "flag_inativo_90d",
        "flag_acessou_hoje", "flag_admin", "flag_externo", "flag_sistema_teste",
        "faixa_inatividade", "faixa_antiguidade", "status_engajamento",
    ]].copy()
    fato = fato.merge(dim_usuario[["id_usuario", "codigo"]], on="codigo", how="left")
    fato.insert(0, "id_usuario", fato.pop("id_usuario"))
    fato["data_carga"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"  fato_acesso_usuario: {fato.shape}")
    print("\n  Distribuição status_engajamento:")
    for status, cnt in fato["status_engajamento"].value_counts().items():
        print(f"    {status:<20} {cnt:>4}")

    # Lookup reutilizado por todos os satélites de autorização
    # Chave: codigo_usuario (string, maiúsculo) → id_usuario (int)
    codigo_para_id = (
        fato[["codigo", "id_usuario"]]
        .drop_duplicates("codigo")
        .rename(columns={"codigo": "codigo_usuario"})
    )

    # ── 7. dim_permissao (ações de sistema) ───────────────────────────────────
    print("\n── 7. dim_permissao ────────────────────────────────────────────────")
    fato_relacao = fato[fato["codigo"] != "ANACLARAAPRENDIZ"]
    dim_permissao = pd.merge(
        df_perm,
        fato_relacao[["nome", "id_usuario"]],
        left_on="usuario", right_on="nome",
        how="left",
    )
    dim_permissao = dim_permissao[dim_permissao["id_usuario"].notna()]
    dim_permissao = dim_permissao[["id_usuario", "nome", "usuario", "acao_id", "tem_permissao"]]
    dim_permissao["id_usuario"] = pd.to_numeric(dim_permissao["id_usuario"], errors="coerce")
    dim_permissao.dropna(subset=["id_usuario", "nome"], inplace=True)
    print(f"  dim_permissao: {dim_permissao.shape}")

    # ── 8. dim_perfil_usuario ─────────────────────────────────────────────────
    print("\n── 8. dim_perfil_usuario ───────────────────────────────────────────")
    dim_perfil_usuario = _enriquecer_com_id(perfil_raw, codigo_para_id, "dim_perfil_usuario")
    print(f"  dim_perfil_usuario: {dim_perfil_usuario.shape}")
    print(f"  Perfis únicos:      {dim_perfil_usuario['perfil_codigo'].nunique()}")

    # ── 8c. dim_permissao_perfil ───────────────────────────────────────────────
    print("\n── 8c. dim_permissao_perfil ────────────────────────────────────────")
    dim_permissao_perfil = perm_perfil_raw.copy()

    # Diagnóstico de cobertura
    perfis_com_dados = dim_permissao_perfil["perfil_codigo"].dropna().nunique()
    acoes_cobertas = dim_permissao_perfil["acao_id"].nunique()
    acessos_true = dim_permissao_perfil["tem_permissao"].sum()
    print(f"  dim_permissao_perfil: {dim_permissao_perfil.shape}")
    print(f"  Perfis com dados:     {perfis_com_dados}")
    print(f"  Ações cobertas:       {acoes_cobertas}")
    print(f"  Acessos concedidos:   {acessos_true:,}")

    # ── 8b. dim_mapeamento_cargo_perfil ──────────────────────────────────────
    print("\n── 8b. dim_mapeamento_cargo_perfil ─────────────────────────────────")
    dim_mapeamento = _gerar_mapeamento_cargo_perfil()

    # Enriquecer com flag de aderência: cruzar perfis do usuário com esperados
    # Grain resultado: 1 linha por usuario × cargo × perfil_esperado
    perfis_usuario = dim_perfil_usuario[["id_usuario", "perfil_codigo"]].copy()
    cargo_usuario = dim_usuario[["id_usuario", "cargo"]].copy()

    aderencia = cargo_usuario.merge(
        dim_mapeamento[dim_mapeamento["mapeamento_definido"]],
        on="cargo", how="left"
    )
    aderencia = aderencia.merge(
        perfis_usuario.rename(columns={"perfil_codigo": "perfil_esperado"}),
        on=["id_usuario", "perfil_esperado"],
        how="left",
        indicator=True,
    )
    # _merge == 'both' → usuário TEM o perfil esperado
    aderencia["tem_perfil_esperado"] = aderencia["_merge"] == "both"
    aderencia = aderencia.drop(columns=["_merge"])

    # Resumo por usuário: tem ao menos 1 perfil esperado?
    resumo_aderencia = (
        aderencia.groupby("id_usuario")["tem_perfil_esperado"]
        .any()
        .reset_index()
        .rename(columns={"tem_perfil_esperado": "tem_ao_menos_um_perfil_esperado"})
    )

    # Usuários com cargo mapeado mas sem nenhum perfil esperado = divergência
    n_diverge = (~resumo_aderencia["tem_ao_menos_um_perfil_esperado"]).sum()
    n_match = resumo_aderencia["tem_ao_menos_um_perfil_esperado"].sum()
    print(f"  dim_mapeamento_cargo_perfil: {dim_mapeamento.shape}")
    print(f"  Cargos com mapeamento definido: "
          f"{dim_mapeamento['mapeamento_definido'].sum()}")
    print(f"  Usuários com ao menos 1 perfil esperado (match): {n_match}")
    print(f"  Usuários com cargo mapeado mas sem perfil esperado (diverge): {n_diverge}")

    # ── 9. fato_permissao_empresa ─────────────────────────────────────────────
    print("\n── 9. fato_permissao_empresa ───────────────────────────────────────")
    fato_permissao_empresa = _enriquecer_com_id(
        empresa_raw, codigo_para_id, "fato_permissao_empresa"
    )
    print(f"  fato_permissao_empresa: {fato_permissao_empresa.shape}")
    print(f"  Empresas distintas:     {fato_permissao_empresa['empresa_codigo'].nunique()}")
    print(f"  Com acesso=True:        {fato_permissao_empresa['tem_acesso'].sum():,}")

    # ── 10. fato_permissao_departamento ───────────────────────────────────────
    print("\n── 10. fato_permissao_departamento ─────────────────────────────────")
    fato_permissao_departamento = _enriquecer_com_id(
        depto_raw, codigo_para_id, "fato_permissao_departamento"
    )
    print(f"  fato_permissao_departamento: {fato_permissao_departamento.shape}")
    print(f"  Departamentos distintos:     "
          f"{fato_permissao_departamento['departamento_codigo'].nunique()}")
    print(f"  Com acesso=True:             "
          f"{fato_permissao_departamento['tem_acesso'].sum():,}")

    # ── 11. fato_permissao_obra ───────────────────────────────────────────────
    print("\n── 11. fato_permissao_obra ─────────────────────────────────────────")
    fato_permissao_obra = _enriquecer_com_id(
        obra_raw, codigo_para_id, "fato_permissao_obra"
    )
    print(f"  fato_permissao_obra: {fato_permissao_obra.shape}")
    print(f"  Obras distintas:     {fato_permissao_obra['obra_codigo'].nunique()}")
    print(f"  Com acesso=True:     {fato_permissao_obra['tem_acesso'].sum():,}")

    # ── 12. Exportação ────────────────────────────────────────────────────────
    print("\n── 12. Exportação ──────────────────────────────────────────────────")
    salvar_tabela(dim_usuario, "dim_usuario", output_dir)
    salvar_tabela(fato, "fato_acesso_usuario", output_dir)
    salvar_tabela(dim_acoes, "dim_acoes_sistema", output_dir)
    salvar_tabela(dim_permissao, "dim_permissao", output_dir)
    salvar_tabela(dim_perfil_usuario, "dim_perfil_usuario", output_dir)
    salvar_tabela(dim_permissao_perfil, "dim_permissao_perfil", output_dir)
    salvar_tabela(dim_mapeamento, "dim_mapeamento_cargo_perfil", output_dir)
    salvar_tabela(aderencia, "dim_aderencia_perfil", output_dir)
    salvar_tabela(fato_permissao_empresa, "fato_permissao_empresa", output_dir)
    salvar_tabela(fato_permissao_departamento, "fato_permissao_departamento", output_dir)
    salvar_tabela(fato_permissao_obra, "fato_permissao_obra", output_dir)

    print("\n── Resumo final ────────────────────────────────────────────────────")
    for nome_tab, tabela in {
        "dim_usuario": dim_usuario,
        "fato_acesso_usuario": fato,
        "dim_acoes_sistema": dim_acoes,
        "dim_permissao": dim_permissao,
        "dim_perfil_usuario": dim_perfil_usuario,
        "dim_permissao_perfil": dim_permissao_perfil,
        "dim_mapeamento_cargo_perfil": dim_mapeamento,
        "dim_aderencia_perfil": aderencia,
        "fato_permissao_empresa": fato_permissao_empresa,
        "fato_permissao_departamento": fato_permissao_departamento,
        "fato_permissao_obra": fato_permissao_obra,
    }.items():
        print(f"  {nome_tab:<34} {str(tabela.shape):>14}")

    print("""
── Relacionamentos Power BI ──────────────────────────────────────────────────
  dim_usuario[id_usuario]  ──1:1──  fato_acesso_usuario[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  dim_permissao[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  dim_perfil_usuario[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  fato_permissao_empresa[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  fato_permissao_departamento[id_usuario]
  dim_usuario[id_usuario]  ──1:N──  fato_permissao_obra[id_usuario]
  dim_acoes_sistema[codigo] ─M:1──  dim_permissao[acao_id]

── Camadas de autorização do SIENGE ─────────────────────────────────────────
  Camada 1 — O QUÊ:  dim_permissao + dim_perfil_usuario
             (ações disponíveis no sistema, herdadas do perfil)
  Camada 2 — ONDE:   fato_permissao_empresa
                     fato_permissao_departamento
                     fato_permissao_obra
             (escopo operacional que restringe onde o usuário pode agir)
  Camada 3 — QUEM:   dim_usuario + fato_acesso_usuario
             (identidade, cargo, engajamento)

  Acesso efetivo = (permissões do perfil) ∩ (escopo autorizado por entidade)
  Ex: usuário com "Lançar NF" no perfil só consegue lançar se a empresa
  ou obra específica também estiver autorizada nas camadas 2.
""")


if __name__ == "__main__":
    import logging

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    executar()

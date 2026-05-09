import pandas as pd
import re
from openpyxl import load_workbook

INPUT_FILE = r'C:\Users\kaua.rodrigo\Documents\etl_sienge\stages\transform\input\financimento_bancario\financiamento.xlsx'
OUTPUT_FILE = r"C:\Users\kaua.rodrigo\Documents\etl_sienge\stages\transform\input\financimento_bancario\financiamento_derretido.csv"

wb = load_workbook(INPUT_FILE, data_only=True)
sheet = wb.active

# Lê todas as linhas como lista de valores
rows = []
for row in sheet.iter_rows(values_only=True):
    rows.append(list(row))


def clean(val):
    if val is None:
        return None
    val = str(val).strip()
    return val if val else None


def parse_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        # Formato BR: "1.234,56" → remove pontos, troca vírgula
        s = str(val).strip()
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        return float(s)
    except:
        return None


# Identificadores de linhas de cabeçalho do contrato
HEADER_KEYS = {
    "Empresa": "empresa",
    "Instituição": "instituicao",
    "Centro de custo": "centro_de_custo",
    "Valor do contrato": "valor_do_contrato",
    "Valor IOF": "valor_iof",
    "Valor IOC": "valor_ioc",
    "Tarifa": "tarifa",
    "Data de emissão": "data_de_emissao",
    "Período das parcelas": "periodo_das_parcelas",
    "Sistema de amortização": "sistema_de_amortizacao",
    "Juros pré-fixado": "juros_pre_fixado",
    "Juros pós-fixado": "juros_pos_fixado",
    "Carência": "carencia",
    "Prazo de carência": "prazo_de_carencia",
    "Título": "titulo",
    "Documento": "documento",
}

contracts = []
current_header = {}
in_parcelas = False

i = 0
while i < len(rows):
    row = rows[i]
    first_cell = clean(row[0]) if row else None

    # Detecta início de um novo contrato pelo campo "Empresa"
    if first_cell == "Empresa":
        # Salva o contrato anterior se existir
        if current_header and current_header.get("_parcelas"):
            contracts.append(current_header)

        current_header = {"_parcelas": []}
        in_parcelas = False

        # Lê as linhas de cabeçalho (4 linhas por bloco, 2 blocos lado a lado)
        # Estrutura: col0=label_esq, col3=valor_esq, col4=label_dir, col7=valor_dir
        header_rows_block = rows[i:i + 7]
        for hrow in header_rows_block:
            # Lado esquerdo: col0 = label, col3 = valor
            lbl_esq = clean(hrow[0]) if len(hrow) > 0 else None
            val_esq_idx = None
            # Encontra primeiro valor não-nulo após col0
            for ci in range(1, min(5, len(hrow))):
                if clean(hrow[ci]) is not None:
                    val_esq_idx = ci
                    break
            val_esq = clean(hrow[val_esq_idx]) if val_esq_idx else None

            # Lado direito: busca label após posição 4
            lbl_dir = None
            val_dir = None
            for ci in range(4, min(len(hrow), 10)):
                if clean(hrow[ci]) is not None:
                    lbl_dir = clean(hrow[ci])
                    # Valor logo após
                    for cj in range(ci + 1, min(len(hrow), 13)):
                        if clean(hrow[cj]) is not None:
                            val_dir = clean(hrow[cj])
                            break
                    break

            if lbl_esq and lbl_esq in HEADER_KEYS:
                current_header[HEADER_KEYS[lbl_esq]] = val_esq
            if lbl_dir and lbl_dir in HEADER_KEYS:
                current_header[HEADER_KEYS[lbl_dir]] = val_dir

        i += 7
        continue

    # Detecta linhas de Título e Documento (ficam isoladas)
    if first_cell in ("Título", "Titulo"):
        val = clean(row[3]) or clean(row[1]) or clean(row[2])
        current_header["titulo"] = val
        i += 1
        continue

    if first_cell == "Documento":
        val = clean(row[3]) or clean(row[1]) or clean(row[2])
        current_header["documento"] = val
        i += 1
        continue

    # Detecta linha de cabeçalho das parcelas
    if first_cell == "Parcela":
        in_parcelas = True
        i += 1
        continue

    # Detecta linha de totais (fim das parcelas)
    if first_cell == "Totais":
        in_parcelas = False
        i += 1
        continue

    # Lê linhas de parcelas
    if in_parcelas and first_cell is not None:
        try:
            num_parcela = int(first_cell)
            parcela = {
                "parcela": num_parcela,
                "data_vencimento": clean(row[2]),
                "amortizacao": parse_float(row[4]),
                "juros_pre": parse_float(row[5]),
                "juros_pos": parse_float(row[6]),
                "valor_da_parcela": parse_float(row[8]),
                "saldo_devedor": parse_float(row[9]),
                "situacao": clean(row[11]),
            }
            current_header["_parcelas"].append(parcela)
        except (ValueError, TypeError):
            pass

    i += 1

# Adiciona o último contrato
if current_header and current_header.get("_parcelas"):
    contracts.append(current_header)

# Derrete: uma linha por parcela com todos os campos do contrato
records = []
for contract in contracts:
    parcelas = contract.pop("_parcelas", [])
    for p in parcelas:
        row_data = {**contract, **p}
        records.append(row_data)

df = pd.DataFrame(records)

# Reordena colunas
col_order = [
    "empresa", "instituicao", "centro_de_custo", "titulo", "documento",
    "valor_do_contrato", "valor_iof", "valor_ioc", "tarifa",
    "data_de_emissao", "periodo_das_parcelas", "sistema_de_amortizacao",
    "juros_pre_fixado", "juros_pos_fixado", "carencia", "prazo_de_carencia",
    "parcela", "data_vencimento", "amortizacao", "juros_pre", "juros_pos",
    "valor_da_parcela", "saldo_devedor", "situacao",
]
col_order = [c for c in col_order if c in df.columns]
df = df[col_order]

print(f"✅ Concluído! {len(df)} linhas geradas de {len(contracts)} contratos.")
print(f"Contratos encontrados: {len(contracts)}")
print(f"Colunas: {list(df.columns)}")
print(df.head(3).to_string())

# ─────────────────────────────────────────────
# 1. LEITURA DOS DADOS
# ─────────────────────────────────────────────

import numpy as np
from pathlib import Path


files = Path('../transform/input/financimento_bancario/consultas_parcelas/').glob('*.csv*')

dfs_csv = []

for file in files:
    df_csv = pd.read_csv(
        file,
        sep=None, engine="python", encoding="utf-8")

    dfs_csv.append(df_csv)

df_csv = pd.concat(dfs_csv, ignore_index=True)


df_csv = df_csv.dropna(subset=["Título", "Parcela"])


# Parse da chave de join: "1/60" → parcela=1
df_csv["_parcela_num"] = df_csv["Parcela"].str.split("/").str[0].astype(int)
df_csv["_titulo"] = df_csv["Título"].astype(int)


# Limpeza de valores monetários do CSV (R$ 255.005,10 → float)
def parse_brl(val):
    if pd.isna(val):
        return np.nan
    s = str(val).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(s)
    except:
        return np.nan


cols_brl = ["Valor bruto", "Valor da baixa", "Saldo em aberto"]
for c in cols_brl:
    df_csv[c + "_num"] = df_csv[c].apply(parse_brl)

# Datas do CSV
df_csv["Data do pagamento_dt"] = pd.to_datetime(df_csv["Data do pagamento"], dayfirst=True, errors="coerce")
df_csv["Data vencimento_dt"] = pd.to_datetime(df_csv["Data vencimento"], dayfirst=True, errors="coerce")

# Campos selecionados do CSV para enriquecer
csv_join = df_csv[[
    "_titulo", "_parcela_num",
    "Indexador",
    "Valor bruto_num", "Valor da baixa_num", "Saldo em aberto_num",
    "Data do pagamento_dt", "Dias de atraso",
    "Status da parcela"
]].rename(columns={
    "_titulo": "titulo",
    "_parcela_num": "parcela",
    "Indexador": "indexador_csv",
    "Valor bruto_num": "valor_bruto_csv",
    "Valor da baixa_num": "valor_baixa_csv",
    "Saldo em aberto_num": "saldo_aberto_csv",
    "Data do pagamento_dt": "data_pagamento_csv",
    "Dias de atraso": "dias_atraso_csv",
    "Status da parcela": "status_parcela_csv",
})

# ─────────────────────────────────────────────
# 2. JOIN
# ─────────────────────────────────────────────
df["titulo"] = df["titulo"].astype(str)
csv_join["titulo"] = csv_join["titulo"].astype(str)

df["parcela"] = df["parcela"].astype(int)
csv_join["parcela"] = csv_join["parcela"].astype(int)


df = df.merge(csv_join, on=["titulo", "parcela"], how="left")

# ─────────────────────────────────────────────
# 3. COLUNAS CALCULADAS POR CONTRATO
# ─────────────────────────────────────────────
df["data_vencimento_dt"] = pd.to_datetime(df["data_vencimento"], dayfirst=True, errors="coerce")
df["data_de_emissao_dt"] = pd.to_datetime(df["data_de_emissao"], dayfirst=True, errors="coerce")

results = []

for titulo, grp in df.groupby("titulo"):
    g = grp.sort_values("parcela").reset_index(drop=True)

    # ── PRAZO EM MESES ──────────────────────────────────────────
    datas = g["data_vencimento_dt"]
    prazo_meses = (datas.max().year - datas.min().year) * 12 + \
                  (datas.max().month - datas.min().month) + 1

    # ── TIPO DE AMORTIZAÇÃO ─────────────────────────────────────
    tipo_amort = str(g["sistema_de_amortizacao"].iloc[0])

    # ── TAXA MENSAL REAL ────────────────────────────────────────
    # O sistema SIENGO armazena juros_pre_fixado como taxa MENSAL decimal
    # ex: 0.0135 = 1.35% a.m. | 0.0069 = 0.69% a.m. | 0.0053 = 0.53% a.m.
    taxa_raw = g["juros_pre_fixado"].iloc[0]
    if pd.notna(taxa_raw) and float(taxa_raw) > 0:
        taxa_mensal = float(taxa_raw)  # já é mensal
        taxa_label = f"{taxa_mensal * 100:.4f}% a.m. / {((1 + taxa_mensal) ** 12 - 1) * 100:.4f}% a.a."
    else:
        taxa_mensal = np.nan
        taxa_label = "CDI / Pós-fixado"

    # ── TAXA ANUAL ───────────────────────────────────────────────
    taxa_anual = (1 + taxa_mensal) ** 12 - 1 if pd.notna(taxa_mensal) else np.nan

    # ── SALDO ANTERIOR (para cálculo de juros e PMT) ──────────────
    g["saldo_anterior"] = g["saldo_devedor"].shift(1)
    g.loc[0, "saldo_anterior"] = float(g["valor_do_contrato"].iloc[0])

    # ── JUROS CALCULADO (taxa × saldo_anterior) ──────────────────
    # Usa o juros_pre se existir; caso contrário recalcula
    if pd.notna(taxa_mensal):
        g["juros_calculado"] = g["saldo_anterior"] * taxa_mensal
    else:
        g["juros_calculado"] = g["juros_pre"]

    # ── PMT = Amortização + Juros ────────────────────────────────
    # No SAC: PMT = amortizacao + juros_calculado
    # (diferente do valor_da_parcela que pode ter IOF/variações)
    g["pmt_calculado"] = g["amortizacao"] + g["juros_calculado"]

    # ── DESVIO ENTRE PARCELAS (dias) ─────────────────────────────
    g["desvio_dias"] = g["data_vencimento_dt"].diff().dt.days

    # ── ATRIBUIÇÕES ──────────────────────────────────────────────
    g["prazo_meses"] = prazo_meses
    g["tipo_amortizacao"] = tipo_amort
    g["taxa_mensal"] = taxa_mensal
    g["taxa_anual"] = taxa_anual
    g["taxa_label"] = taxa_label

    results.append(g)

df = pd.concat(results).reset_index(drop=True)

# ─────────────────────────────────────────────
# 4. ORGANIZAR COLUNAS FINAIS
# ─────────────────────────────────────────────
col_order = [
    # Identificação
    "empresa", "instituicao", "centro_de_custo", "titulo", "documento",
    # Características do contrato
    "valor_do_contrato", "valor_iof", "valor_ioc", "tarifa",
    "data_de_emissao",
    "prazo_meses",
    "periodo_das_parcelas",
    "tipo_amortizacao",
    "taxa_mensal", "taxa_anual", "taxa_label",
    "juros_pre_fixado", "juros_pos_fixado",
    "carencia", "prazo_de_carencia",
    # Parcela
    "parcela", "data_vencimento", "desvio_dias",
    "saldo_anterior",
    "amortizacao",
    "juros_calculado",
    "pmt_calculado",
    "valor_da_parcela",
    "saldo_devedor",
    "situacao",
    # Dados do CSV
    "indexador_csv", "valor_bruto_csv", "valor_baixa_csv",
    "saldo_aberto_csv", "data_pagamento_csv", "dias_atraso_csv", "status_parcela_csv",
]
col_order = [c for c in col_order if c in df.columns]
df_out = df[col_order].copy()

# ─────────────────────────────────────────────
# 5. SALVAR COM FORMATAÇÃO
# ─────────────────────────────────────────────
df.to_csv(OUTPUT_FILE, index=False, sep=';', decimal=',')

# ─────────────────────────────────────────────
# 6. SUMÁRIO
# ─────────────────────────────────────────────
print(f"\n✅ Gerado: {OUTPUT_FILE}")
print(f"   Linhas: {len(df_out)} | Colunas: {len(df_out.columns)}")
print(f"\n{'─' * 60}")
print(f"{'Título':<10} {'Prazo':>8} {'Taxa Mensal':>12} {'Taxa Anual':>12} {'Tipo Amort':<12} {'Taxa Label'}")
print(f"{'─' * 60}")
for titulo, grp in df_out.groupby("titulo"):
    r = grp.iloc[0]
    print(
        f"{titulo:<10} {r['prazo_meses']:>7}m {r['taxa_mensal'] * 100:>11.4f}% {r['taxa_anual'] * 100:>11.4f}%  {r['tipo_amortizacao']:<12} {r['taxa_label']}")
print(f"{'─' * 60}")
print(f"\nNOVAS COLUNAS adicionadas:")
novas = ["prazo_meses", "tipo_amortizacao", "taxa_mensal", "taxa_anual", "taxa_label",
         "saldo_anterior", "juros_calculado", "pmt_calculado", "desvio_dias",
         "indexador_csv", "valor_bruto_csv", "valor_baixa_csv", "saldo_aberto_csv",
         "data_pagamento_csv", "dias_atraso_csv", "status_parcela_csv"]
for c in novas:
    if c in df_out.columns:
        print(f"  + {c}")

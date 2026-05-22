"""
stages/transform/transform_contas_recebidas.py
═══════════════════════════════════════════════
Transform unificado – Contas Recebidas

Uso
───
  python transform_contas_recebidas.py            → consolidar + fechamento
  python transform_contas_recebidas.py --so-consolidar  → só lê brutos e salva consolidado

Saída por centro de custo
─────────────────────────
  dados_consolidados/
    {slug}_consolidado_sintetico.xlsx   ← 2 abas: Consolidado | Fechamento
    {slug}_consolidado_analitico.xlsx   ← 2 abas: Analítico   | Fechamento
    {slug}_acompanhamento.xlsx          ← pivot mensal por cliente

  Aba Acompanhamento
    Série histórica pivotada por mês (dt_baixa → colunas).
    Configurável via COLS_ACOMPANHAMENTO_ANTES e COLS_ACOMPANHAMENTO_DEPOIS.

Consolidado geral (raiz de contas_recebidas/)
  consolidado_geral_analitico.csv
  Concatenação de todos os analíticos — dado granular para Power BI.
"""

from __future__ import annotations

import argparse
import logging
import re

from datetime import date, timedelta, datetime
from pathlib import Path
import locale

import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Paths ─────────────────────────────────────────────────────────────────────

AUXILIAR_PATH = (
        Path(__file__).resolve().parents[2]
        / "stages" / "extract" / "reference" / "auxiliar_contas_recebidas.xlsx"
)

BASE_INPUT_DIR = (
        Path(__file__).resolve().parents[2]
        / "stages" / "transform" / "input" / "contas_recebidas"
)

BASE_OUTPUT_DIR = (
        Path(__file__).resolve().parents[2]
        / "stages" / "transform" / "output" / "contas_recebidas"
)

EXTRATO_PATH = (
        Path(__file__).resolve().parents[2]
        / "stages" / "extract" / "reference" / "extrato_empreendimentos.xls"
)

# ── Configuração da aba Acompanhamento ────────────────────────────────────────
#
# Liste aqui as colunas do df_sin_fechamento que devem aparecer
# ANTES das colunas mensais pivotadas (na ordem desejada).
# Colunas inválidas (inexistentes no df) são ignoradas silenciosamente.

COLS_ACOMPANHAMENTO_ANTES: list[str] = [
    "cliente",  # normalizado para "cliente" (sem _x/_y) antes do pivot
    "titulo",
    "novo_cliente",
    "vgv_vendido",
    "pct_repasse",
    "total_por_cliente"

]

# Colunas que aparecem DEPOIS das colunas mensais pivotadas.
# Métricas vindas de contas_a_receber (inadimplente, a_vencer, carteira_total, vgv_vendido)
# são trazidas via join depois do pivot — não somam, usam "first".
# contrato_total e check são calculados automaticamente após o pivot.

COLS_ACOMPANHAMENTO_DEPOIS: list[str] = [
    # ── Calculados após o pivot (soma das colunas mensais) ─────────────────
    "valor_liquido_total",  # soma histórica de todos os *_liquido
    "pct_repasse",  # percentual de repasse do centro
    "valor_liquido_repasse",  # soma histórica de todos os *_repasse
    "valor_a_repassar",

    # ── Métricas de carteira (vêm de contas_a_receber via merge) ──────────
    "carteira_total",  # inadimplente + a_vencer
    "inadimplente",  # saldo vencido antes do mês fechado
    "a_vencer",  # saldo a vencer a partir do mês fechado

    # ── Derivados (calculados sobre o df achatado) ─────────────────────────
    "contrato_total",  # valor_liquido_total + carteira_total
    "check",  # contrato_total - vgv_vendido
]

# Colunas de valor que serão pivotadas por mês (dt_baixa).
# Para cada coluna aqui, serão geradas colunas no formato "jan/25_liquido", etc.

COLS_PIVOT_VALORES: list[str] = [
    "liquido",
    "bruto",
]

# ── Mapeamento de meses para PT-BR ────────────────────────────────────────────

MESES_PTBR = {
    1: "jan", 2: "fev", 3: "mar", 4: "abr",
    5: "mai", 6: "jun", 7: "jul", 8: "ago",
    9: "set", 10: "out", 11: "nov", 12: "dez",
}

ORDEM_MESES = {v: k for k, v in MESES_PTBR.items()}

# ── Estilos CLEAN — única paleta usada em todas as abas ──────────────────────

_THIN_CL = Side(style="thin", color="D0D0D0")
_MED_CL = Side(style="medium", color="505050")
_BORDA_CL = Border(left=_THIN_CL, right=_THIN_CL, top=_THIN_CL, bottom=_THIN_CL)
_BORDA_MED_CL = Border(left=_MED_CL, right=_MED_CL, top=_MED_CL, bottom=_MED_CL)

_HDR_FONT_CL = Font(bold=True, color="FFFFFF", size=10)
_HDR_FILL_CL = PatternFill("solid", start_color="404040")  # cabeçalho padrão
_TITLE_FILL_CL = PatternFill("solid", start_color="1A1A1A")  # título principal
_ALT_FILL_CL = PatternFill("solid", start_color="F5F5F5")  # zebra claro (linhas pares)
_TOTAL_FILL_CL = PatternFill("solid", start_color="2B2B2B")  # linha total
_SUBTOT_FILL_CL = PatternFill("solid", start_color="DCDCDC")  # subtotal por cliente
_SECT_FILL_CL = PatternFill("solid", start_color="EBEBEB")  # seção
_PIVOT_HDR_CL = PatternFill("solid", start_color="595959")  # cabeçalho de mês pivotado
_PIVOT_ALT_CL = PatternFill("solid", start_color="FAFAFA")  # zebra alternada pivot
_NOVO_FILL_CL = PatternFill("solid", start_color="F0F0F0")  # destaque novo cliente (cinza suave)

# Atalhos mantidos para compatibilidade com referências espalhadas no código
_BORDA = _BORDA_CL
_BORDA_MED = _BORDA_MED_CL
_HDR_FONT = _HDR_FONT_CL
_HDR_FILL = _HDR_FILL_CL
_TITLE_FILL = _TITLE_FILL_CL
_ALT_FILL = _ALT_FILL_CL
_TOTAL_FILL = _TOTAL_FILL_CL
_SUBTOT_FILL = _SUBTOT_FILL_CL
_NOVO_FILL = _NOVO_FILL_CL
_NOVO_TOT = PatternFill("solid", start_color="C8C8C8")
_SECT_FILL = _SECT_FILL_CL

# ── Regex / constantes ────────────────────────────────────────────────────────

_RE_TS = re.compile(r"^\d{2}/\d{2}/\d{4} - \d{2}:\d{2}:\d{2}$")
_RE_COD = re.compile(r"\s*\(\d+\)\s*$")
RODAPE = ("Total geral", "Total da empresa", "Total de parcelas",
          "Total de títulos", "(*)", "(C)", "(S)", "SIENGE")

# ── Mapeamentos de colunas ────────────────────────────────────────────────────

COLUNAS_SINTETICO_JUROS = {
    0: "cliente", 5: "amortizacao", 7: "juros", 8: "correcao",
    9: "acrescimo", 11: "seguro", 12: "taxa_adm", 13: "desconto", 14: "liquido",
}
COLUNAS_SINTETICO_PADRAO = {
    0: "cliente", 5: "vl_baixa", 7: "acrescimo", 8: "seguro",
    10: "taxa_adm", 11: "desconto", 12: "liquido",
}
COLUNAS_ANALITICO_JUROS = {
    0: "dt_baixa", 1: "cliente", 4: "documento", 7: "titulo",
    8: "parcela", 9: "tc", 10: "unidade_principal", 12: "portador",
    13: "operacao", 14: "data_vencimento", 15: "amortizacao", 16: "juros",
    18: "correcao", 19: "acrescimo", 20: "seguro", 21: "taxa_adm",
    22: "desconto", 23: "liquido",
}
COLUNAS_ANALITICO_PADRAO = {
    0: "dt_baixa", 1: "cliente", 4: "dt_emissao", 5: "documento",
    7: "titulo", 8: "parcela", 9: "tc", 10: "unidade_principal",
    11: "portador", 12: "operacao", 13: "data_vencimento", 14: "vl_baixa",
    15: "acrescimo", 16: "seguro", 17: "taxa_adm", 18: "desconto", 19: "liquido",
}

COLUNAS_NUM_JUROS = ["amortizacao", "juros", "correcao", "acrescimo",
                     "seguro", "taxa_adm", "desconto", "liquido"]
COLUNAS_NUM_PADRAO = ["vl_baixa", "acrescimo", "seguro",
                      "taxa_adm", "desconto", "liquido"]
LAYOUT = {
    "juros embutidos": (COLUNAS_SINTETICO_JUROS, COLUNAS_ANALITICO_JUROS, COLUNAS_NUM_JUROS),
    "padrão": (COLUNAS_SINTETICO_PADRAO, COLUNAS_ANALITICO_PADRAO, COLUNAS_NUM_PADRAO),
}


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS GERAIS
# ═════════════════════════════════════════════════════════════════════════════

def _slug(nome: str) -> str:
    return nome.strip().lower().replace(" ", "_").replace("-", "")[:40]


def _nome_sienge(c: str) -> str:
    return _RE_COD.sub("", str(c)).strip().upper()


def _eh_rodape(v) -> bool:
    s = str(v).strip()
    return s.startswith(RODAPE) or bool(_RE_TS.match(s))


def _mes_anterior() -> tuple[date, date, str]:
    hoje = date.today()
    ini = (hoje.replace(day=1) - timedelta(days=1)).replace(day=1)
    fim = ini + relativedelta(months=1) - timedelta(days=1)
    return ini, fim, ini.strftime("%m/%Y")


def _layout(centro: dict) -> tuple[dict, dict, list[str]]:
    tipo = str(centro.get("tipo_coluna", "")).strip().lower()
    return LAYOUT.get(tipo, LAYOUT["juros embutidos"])


def _inicio_dados(centro: dict) -> int:
    return 10 if str(centro.get("com_centro", "sim")).strip().lower() == "sim" else 9


# ── Formatação de célula ──────────────────────────────────────────────────────

def _fmt_valor(cell):
    cell.number_format = "#,##0.00"
    cell.alignment = Alignment(horizontal="right")


def _fmt_pct(cell):
    cell.number_format = "0.00%"
    cell.alignment = Alignment(horizontal="center")


def _fmt_data(cell):
    cell.number_format = "dd/mm/yyyy"
    cell.alignment = Alignment(horizontal="center")


def _aplicar_celula(cell, col: str, colunas_valor: set, colunas_data: set,
                    e_novo: bool = False):
    cell.border = _BORDA_CL
    cell.font = Font(size=10, color="1A1A1A")
    if col in colunas_data:
        _fmt_data(cell)
    elif col == "pct_repasse":
        _fmt_pct(cell)
    elif col in colunas_valor:
        _fmt_valor(cell)
    elif col == "cliente":
        cell.alignment = Alignment(horizontal="left")
    elif col == "novo_cliente":
        cell.alignment = Alignment(horizontal="center")
        if e_novo:
            cell.font = Font(size=10, bold=True, color="1A1A1A")
    else:
        cell.alignment = Alignment(horizontal="center")


# ── Helpers de layout de worksheet ───────────────────────────────────────────

def _titulo_ws(ws, texto: str, n_cols: int, row: int = 1) -> None:
    end = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = _TITLE_FILL_CL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _meta_ws(ws, texto: str, n_cols: int, row: int = 2) -> None:
    end = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(italic=True, size=10, color="767676")
    c.alignment = Alignment(horizontal="left", vertical="center")


def _cabecalhos_ws(ws, cabecalhos: list[str], row: int = 3) -> None:
    for ci, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=row, column=ci, value=cab)
        c.font = _HDR_FONT_CL
        c.fill = _HDR_FILL_CL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDA_CL
    ws.row_dimensions[row].height = 30


def _secao_ws(ws, texto: str, n_cols: int, row: int,
              fill=None, cor: str = "F5F5F5") -> None:
    fill = fill or _SECT_FILL_CL
    end = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, size=11, color="2B2B2B")
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _BORDA_CL
    ws.row_dimensions[row].height = 18


def _linha_total(ws, ri: int, n_cols: int, label: str,
                 df: pd.DataFrame, colunas_df: list[str],
                 colunas_valor: set, pct: float,
                 fill=None, cor_fonte: str = "FFFFFF") -> None:
    fill = fill or _TOTAL_FILL_CL
    tf = Font(bold=True, color=cor_fonte, size=10)
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=ri, column=ci)
        c.fill = fill
        c.font = tf
        c.border = _BORDA_MED_CL
    ws.cell(row=ri, column=1, value=label).alignment = Alignment(horizontal="center")
    for ci, col in enumerate(colunas_df, 1):
        if col in colunas_valor and col in df.columns:
            c = ws.cell(row=ri, column=ci, value=df[col].sum())
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        elif col == "pct_repasse":
            c = ws.cell(row=ri, column=ci, value=pct)
            c.number_format = "0.00%"
            c.alignment = Alignment(horizontal="center")


# ── Helpers de layout CLEAN (usados apenas no acompanhamento) ─────────────────

def _titulo_ws_cl(ws, texto: str, n_cols: int, row: int = 1) -> None:
    end = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = _TITLE_FILL_CL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _meta_ws_cl(ws, texto: str, n_cols: int, row: int = 2) -> None:
    end = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(italic=True, size=10, color="767676")
    c.alignment = Alignment(horizontal="left", vertical="center")


def _cabecalho_cl(ws, col_idx: int, row: int, texto: str,
                  fill: PatternFill = None) -> None:
    """Aplica estilo de cabeçalho clean a uma célula específica."""
    fill = fill or _HDR_FILL_CL
    c = ws.cell(row=row, column=col_idx, value=texto)
    c.font = _HDR_FONT_CL
    c.fill = fill
    c.border = _BORDA_CL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _linha_total_cl(ws, ri: int, n_cols: int, label: str,
                    df: pd.DataFrame, colunas_df: list[str],
                    colunas_valor: set) -> None:
    """Linha de total no estilo clean."""
    tf = Font(bold=True, color="FFFFFF", size=10)
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=ri, column=ci)
        c.fill = _TOTAL_FILL_CL
        c.font = tf
        c.border = _BORDA_MED_CL
    ws.cell(row=ri, column=1, value=label).alignment = Alignment(horizontal="left")
    for ci, col in enumerate(colunas_df, 1):
        if col in colunas_valor and col in df.columns:

            # percentual não soma
            if col == "pct_repasse":
                valor = df[col].iloc[0] if not df.empty else 0
            else:
                valor = df[col].sum()

            c = ws.cell(row=ri, column=ci, value=valor)

            c.number_format = "0.00%" if col == "pct_repasse" else "#,##0.00"
            c.alignment = Alignment(horizontal="right")
            c.font = tf


# ═════════════════════════════════════════════════════════════════════════════
# LEITURA DO AUXILIAR E EXTRATO
# ═════════════════════════════════════════════════════════════════════════════

def carregar_centros_ativos() -> list[dict]:
    df = pd.read_excel(AUXILIAR_PATH, sheet_name="centros_custo")
    df = df[df["ativo"].str.strip().str.lower() == "sim"]
    return df.to_dict(orient="records")


def carregar_extrato(nome_empreendimento: str) -> set[str]:
    if not EXTRATO_PATH.exists():
        logger.warning("Extrato não encontrado → novo_cliente sempre 'N'")
        return set()
    df = pd.read_excel(EXTRATO_PATH, sheet_name="UNIDADES", header=1)
    df.columns = ["cnpj", "empreendimento", "contrato",
                  "nome_mutuario", "cpf", "dt_assinatura", "dt_cri"]
    df["dt_cri"] = pd.to_datetime(df["dt_cri"], errors="coerce")
    df["nome_up"] = df["nome_mutuario"].astype(str).str.strip().str.upper()
    chave = re.escape(nome_empreendimento.upper()[:15])
    df = df[df["empreendimento"].astype(str).str.upper().str.contains(chave, regex=True)]
    ini, fim, label = _mes_anterior()
    mask = (df["dt_cri"].dt.date >= ini) & (df["dt_cri"].dt.date <= fim)
    novos = set(df.loc[mask, "nome_up"].tolist())
    logger.info("Extrato [%s]: %d novo(s) em %s", nome_empreendimento, len(novos), label)
    return novos


# ═════════════════════════════════════════════════════════════════════════════
# LEITURA DOS BRUTOS
# ═════════════════════════════════════════════════════════════════════════════

def _extrair_periodo(df_raw: pd.DataFrame) -> str:
    try:
        m = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(df_raw.iloc[6, 2]))
        if m:
            return f"{m.group(3)}{m.group(2)}"
    except Exception:
        pass
    return "000000"


def ler_vgv_clientes() -> pd.DataFrame:
    df = pd.read_excel('../extract/reference/vgv_clientes.xlsx', decimal=',')
    df.dropna(subset='id título', inplace=True)
    df['id título'] = pd.to_numeric(df['id título'], errors='coerce')

    df.rename(columns={'id título': 'titulo', 'Cliente': 'cliente'}, inplace=True)

    return df[['titulo', 'VGV Vendido', 'cliente']]


def ler_contas_a_receber() -> pd.DataFrame:
    files = Path(r'C:\Users\kaua.rodrigo\Documents\etl_sienge\stages\transform\input\contas_a_receber').glob("*.xlsx*")
    arquivos = []

    for file in files:
        df_raw = pd.read_excel(file, sheet_name="Relatório", header=None)

        colunas_conta_a_receber = {
            0: "data_vecto", 1: "cliente",
            4: "documento", 5: "documento",
            7: "titulo", 9: "parcela",
            10: "tc", 11: "unidade_principal",
            13: "valor_original", 14: "id",
            15: "data_calculo", 16: "saldo_atual",
            19: "dias", 20: "acrescimo",
            21: "desconto", 22: "seguro",
            23: "taxa_adm", 24: "total",
        }
        colunas_num = ["valor_original", "saldo_atual", "acrescimo",
                       "desconto", "seguro", "taxa_adm", "total"]

        dados = df_raw.iloc[9:][list(colunas_conta_a_receber.keys())].copy()
        dados.columns = list(colunas_conta_a_receber.values())
        dados = dados[dados["cliente"].notna()]
        dados = dados[~dados["cliente"].astype(str).str.strip().eq("")]
        dados = dados[~dados["cliente"].astype(str).apply(_eh_rodape)]

        dados["data_vecto"] = pd.to_datetime(
            dados["data_vecto"],
            format="%d/%m/%Y",
            errors="coerce"
        )

        for col in colunas_num:
            dados[col] = pd.to_numeric(dados[col], errors="coerce").fillna(0.0)

        arquivos.append(dados)

    dados = pd.concat(arquivos, ignore_index=True)
    ini_mes_fechado, _, _ = _mes_anterior()
    data_corte = pd.Timestamp(ini_mes_fechado)

    dados["inadimplente"] = np.where(dados["data_vecto"] < data_corte, dados["saldo_atual"], 0.0)
    dados["a_vencer"] = np.where(dados["data_vecto"] >= data_corte, dados["saldo_atual"], 0.0)
    dados["carteira_total"] = dados["inadimplente"] + dados["a_vencer"]

    dados = (
        dados
        .groupby(["cliente", "titulo"], as_index=False)
        .agg({
            **{c: "sum" for c in (colunas_num + ["inadimplente", "a_vencer", "carteira_total"])
               if c in dados.columns}
        })
        .sort_values("cliente")
        .reset_index(drop=True)
    )

    df_vgv = ler_vgv_clientes()

    dados["titulo"] = pd.to_numeric(dados["titulo"], errors="coerce")
    df_vgv["titulo"] = pd.to_numeric(df_vgv["titulo"], errors="coerce")


    dados = dados.merge(df_vgv, how="outer", left_on="titulo", right_on="titulo")

    dados["cliente"] = dados["cliente_x"].combine_first(dados["cliente_y"])
    dados.drop(columns=["cliente_x", "cliente_y"], inplace=True)

    dados.rename(columns={"VGV Vendido": "vgv_vendido"}, inplace=True)

    return dados.reset_index(drop=True)


def ler_analitico(caminho: Path, novos: set[str], centro: dict) -> pd.DataFrame:
    _, cols_ana, colunas_num = _layout(centro)
    df_raw = pd.read_excel(caminho, sheet_name="Relatório", header=None)
    periodo = _extrair_periodo(df_raw)
    inicio = _inicio_dados(centro)

    dados = df_raw.iloc[inicio:][list(cols_ana.keys())].copy()
    dados.columns = list(cols_ana.values())
    dados = dados[dados["cliente"].notna()]
    dados = dados[~dados["cliente"].astype(str).str.strip().eq("")]
    dados = dados[~dados["cliente"].astype(str).str.startswith("Total do cliente")]
    dados = dados[~dados["cliente"].astype(str).apply(_eh_rodape)]

    for col in colunas_num:
        if col == "amortizacao":
            dados[col] = (dados[col].astype(str)
                          .str.replace(" P", "", regex=False).str.strip())
            mask = dados[col].str.contains(",", na=False)
            dados.loc[mask, col] = (dados.loc[mask, col]
                                    .str.replace(".", "", regex=False)
                                    .str.replace(",", ".", regex=False))
        dados[col] = pd.to_numeric(dados[col], errors="coerce").fillna(0.0)

    if "liquido" not in dados.columns and "vl_baixa" in dados.columns:
        dados["liquido"] = dados["vl_baixa"]

    dados["dt_baixa"] = pd.to_datetime(dados["dt_baixa"], format="%d/%m/%Y", errors="coerce")
    dados.insert(0, "periodo", periodo)
    dados["novo_cliente"] = dados["cliente"].apply(
        lambda c: "S" if _nome_sienge(c) in novos else "N"
    )
    return dados.reset_index(drop=True)


# ═════════════════════════════════════════════════════════════════════════════
# CONSOLIDAR – SINTÉTICO / ANALÍTICO
# ═════════════════════════════════════════════════════════════════════════════

def _colunas_excel_sintetico(centro: dict) -> tuple[list[str], list[str], set]:
    _, _, colunas_num = _layout(centro)
    possui_juros = "juros" in colunas_num
    if possui_juros:
        cols_num_exib = ["amortizacao", "juros", "correcao",
                         "acrescimo", "seguro", "taxa_adm", "desconto", "liquido"]
        cabs_num = ["Amortização", "Juros", "Correção",
                    "Acréscimo", "Seguro", "Taxa adm", "Desconto", "Líquido"]
    else:
        cols_num_exib = ["vl_baixa", "acrescimo", "seguro",
                         "taxa_adm", "desconto", "liquido"]
        cabs_num = ["Vl. Baixa", "Acréscimo", "Seguro",
                    "Taxa adm", "Desconto", "Líquido"]

    cabecalhos = ["Cliente", "Novo?"] + cabs_num + ["% Repasse", "Vlr. Líq. Repasse"]
    colunas_df = ["cliente", "novo_cliente"] + cols_num_exib + ["pct_repasse", "valor_liquido_repasse"]
    return cabecalhos, colunas_df, set(cols_num_exib + ["valor_liquido_repasse"])


def _colunas_excel_analitico(centro: dict) -> tuple[list[str], list[str], set]:
    _, _, colunas_num = _layout(centro)
    possui_juros = "juros" in colunas_num
    cols_fixas = ["dt_baixa", "cliente", "novo_cliente", "documento", "titulo",
                  "parcela", "tc", "unidade_principal", "portador",
                  "operacao", "data_vencimento"]
    cabs_fixas = ["Dt. Baixa", "Cliente", "Novo?", "Documento", "Título",
                  "Parcela", "TC", "Unidade", "Portador", "Operação", "Vencimento"]
    if possui_juros:
        cols_num_exib = ["amortizacao", "juros", "correcao",
                         "acrescimo", "seguro", "taxa_adm", "desconto", "liquido"]
        cabs_num = ["Amortização", "Juros", "Correção",
                    "Acréscimo", "Seguro", "Taxa Adm", "Desconto", "Líquido"]
    else:
        cols_num_exib = ["vl_baixa", "acrescimo", "seguro",
                         "taxa_adm", "desconto", "liquido"]
        cabs_num = ["Vl. Baixa", "Acréscimo", "Seguro",
                    "Taxa Adm", "Desconto", "Líquido"]
    cabecalhos = cabs_fixas + cabs_num + ["% Repasse", "Vlr. Líq. Repasse"]
    colunas_df = cols_fixas + cols_num_exib + ["pct_repasse", "valor_liquido_repasse"]
    return cabecalhos, colunas_df, set(cols_num_exib + ["valor_liquido_repasse"])


def _salvar_sintetico(df: pd.DataFrame, ws, nome_cc: str, pct: float, centro: dict) -> None:
    cabecalhos, colunas_df, colunas_valor = _colunas_excel_sintetico(centro)

    cols_remover = {"inadimplente", "a_vencer", "carteira_total", "vgv_vendido", "contrato_total", "check"}
    colunas_df = [c for c in colunas_df if c not in cols_remover]
    colunas_valor = [c for c in colunas_valor if c not in cols_remover]
    cabecalhos = [h for h, c in zip(cabecalhos, _colunas_excel_sintetico(centro)[1])
                  if c not in cols_remover]

    n = len(cabecalhos)
    _titulo_ws(ws, f"Contas Recebidas – {nome_cc}", n, 1)
    _meta_ws(ws, (
        f"% de repasse: {pct:.2%}   |   "
        f"Clientes únicos: {df['cliente'].nunique()}   |   "
        f"Novos (CRI mês ant.): {(df['novo_cliente'] == 'S').sum()}   |   "
        f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    ), n, 2)
    _cabecalhos_ws(ws, cabecalhos, 3)

    for ri, row in enumerate(df.itertuples(index=False), 4):
        e_novo = getattr(row, "novo_cliente") == "S"
        fill = _NOVO_FILL if e_novo else (_ALT_FILL if ri % 2 == 0 else PatternFill())
        for ci, col in enumerate(colunas_df, 1):
            val = getattr(row, col)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            _aplicar_celula(cell, col, colunas_valor, set(), e_novo)

    tr = len(df) + 4
    _linha_total(ws, tr, n, "TOTAL", df, colunas_df, colunas_valor, pct)

    larguras = {"cliente": 42, "novo_cliente": 8, "pct_repasse": 12, "valor_liquido_repasse": 18}
    _, _, colunas_num = _layout(centro)
    for c in colunas_num:
        larguras[c] = 14
    for i, col in enumerate(colunas_df, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 12)
    ws.freeze_panes = "A4"


def _salvar_analitico(df: pd.DataFrame, ws, nome_cc: str, pct: float, centro: dict) -> None:
    cabecalhos, colunas_df, colunas_valor = _colunas_excel_analitico(centro)
    colunas_data = {"dt_baixa", "data_vencimento"}
    n = len(cabecalhos)

    _titulo_ws(ws, f"Contas Recebidas Analítico – {nome_cc}", n, 1)
    _meta_ws(ws, (
        f"% de repasse: {pct:.2%}   |   "
        f"Registros: {len(df)}   |   "
        f"Clientes únicos: {df['cliente'].nunique()}   |   "
        f"Novos (CRI mês ant.): {(df['novo_cliente'] == 'S').sum()}   |   "
        f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    ), n, 2)
    _cabecalhos_ws(ws, cabecalhos, 3)

    ri = 4
    for cliente, grupo in df.groupby("cliente", sort=True):
        e_novo = (grupo["novo_cliente"] == "S").any()
        for row in grupo.itertuples(index=False):
            fill = _NOVO_FILL if e_novo else (_ALT_FILL if ri % 2 == 0 else PatternFill())
            for ci, col in enumerate(colunas_df, 1):
                val = getattr(row, col)
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                _aplicar_celula(cell, col, colunas_valor, colunas_data, e_novo)
            ri += 1

        sf = Font(bold=True, size=10, color="404040")
        for ci in range(1, n + 1):
            c = ws.cell(row=ri, column=ci)
            c.fill = _SUBTOT_FILL
            c.font = sf
            c.border = _BORDA
        ws.cell(row=ri, column=2, value=f"Subtotal – {cliente}").alignment = \
            Alignment(horizontal="left")
        ws.cell(row=ri, column=3, value="S" if e_novo else "N").alignment = \
            Alignment(horizontal="center")
        for ci, col in enumerate(colunas_df, 1):
            if col in colunas_valor and col in grupo.columns:
                c = ws.cell(row=ri, column=ci, value=grupo[col].sum())
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            elif col == "pct_repasse":
                c = ws.cell(row=ri, column=ci, value=pct)
                c.number_format = "0.00%"
                c.alignment = Alignment(horizontal="center")
        ri += 1

    _linha_total(ws, ri, n, "TOTAL GERAL", df, colunas_df, colunas_valor, pct)

    larguras = {"dt_baixa": 12, "cliente": 40, "novo_cliente": 7,
                "documento": 16, "titulo": 14, "parcela": 10, "tc": 8,
                "unidade_principal": 14, "portador": 12, "operacao": 12,
                "data_vencimento": 12, "pct_repasse": 12, "valor_liquido_repasse": 18}
    _, _, colunas_num = _layout(centro)
    for c in colunas_num:
        larguras[c] = 14
    for i, col in enumerate(colunas_df, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 12)
    ws.freeze_panes = "A4"


# ═════════════════════════════════════════════════════════════════════════════
# ABA ACOMPANHAMENTO  (pivot mensal, estilo clean)
# ═════════════════════════════════════════════════════════════════════════════

def _construir_pivot_mensal(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Pivota df_sin_fechamento de linhas mensais para colunas mensais.

    Retorna
    ───────
    (df_pivot, col_map) onde:
      df_pivot  — 1 linha por (cliente, titulo); colunas internas sem '/'
                  ex: "jan_25_liquido"  →  seguro para itertuples
      col_map   — {"jan_25_liquido": "jan/25_liquido"}  para display no Excel

    Por que sem '/'?
    ────────────────
    itertuples() converte '/' em '_' nos atributos do namedtuple,
    então getattr(row, "jan/25_liquido") retorna None silenciosamente.
    Nomes internos sem '/' eliminam esse bug.

    NÃO traz métricas CR — responsabilidade de _salvar_acompanhamento.
    """

    if df.empty or "dt_baixa" not in df.columns:
        return pd.DataFrame(), {}

    df = df.copy()

    # ── Rótulo de mês ─────────────────────────────────────────────────────────
    df["mes_ref"] = (
            df["dt_baixa"].dt.month.map(MESES_PTBR)
            + "/"
            + df["dt_baixa"].dt.strftime("%y")
    )

    # ── Colunas de valor a pivotar ────────────────────────────────────────────
    cols_valor_pivot = [c for c in COLS_PIVOT_VALORES if c in df.columns]
    if not cols_valor_pivot:
        logger.warning("Nenhuma coluna de pivot encontrada; acompanhamento vazio.")
        return pd.DataFrame(), {}

    # ── Índice: COLS_ACOMPANHAMENTO_ANTES apenas (sem métricas CR) ────────────
    antes_validas = [c for c in COLS_ACOMPANHAMENTO_ANTES if c in df.columns]
    if not antes_validas:
        logger.warning("Nenhuma coluna índice para pivot; acompanhamento vazio.")
        return pd.DataFrame(), {}

    # ── Agrega por (índice, mês) ──────────────────────────────────────────────
    df_agg = (
        df
        .groupby(antes_validas + ["mes_ref"], as_index=False)[cols_valor_pivot]
        .sum()
    )

    # ── Pivot ─────────────────────────────────────────────────────────────────
    df_pivot = df_agg.pivot_table(
        index=antes_validas,
        columns="mes_ref",
        values=cols_valor_pivot,
        aggfunc="sum",
        fill_value=0,
    )

    # ── Flatten: (metrica, mes) → col_interna sem '/', col_display com '/' ───
    # MultiIndex vem como ('liquido', 'jan/25') após pivot_table com values=lista
    col_map: dict[str, str] = {}
    new_cols: list[str] = []
    for metrica, mes in df_pivot.columns.to_flat_index():
        col_interna = f"{mes.replace('/', '_')}_{metrica}"  # "jan_25_liquido"
        col_display = f"{mes}_{metrica}"  # "jan/25_liquido"
        col_map[col_interna] = col_display
        new_cols.append(col_interna)

    df_pivot.columns = new_cols
    df_pivot = df_pivot.reset_index()

    # ── Ordena colunas cronologicamente ───────────────────────────────────────
    # "jan_25_liquido" → (25, 1, "liquido")
    def _chave_col(col: str) -> tuple:
        try:
            parts = col.split("_")  # ["jan", "25", "liquido"]
            nome_mes = parts[0]
            ano = parts[1]
            metrica = "_".join(parts[2:])
            return (int(ano), ORDEM_MESES.get(nome_mes, 99), metrica)
        except Exception:
            return (9999, 99, col)

    cols_pivot_sorted = sorted(
        [c for c in df_pivot.columns if c not in antes_validas],
        key=_chave_col,
    )

    colunas_finais = [c for c in antes_validas if c in df_pivot.columns] + cols_pivot_sorted

    df_pivot = (
        df_pivot[colunas_finais]
        .sort_values([c for c in ["cliente", "titulo"] if c in df_pivot.columns])
        .reset_index(drop=True)
    )

    return df_pivot, col_map


def _salvar_acompanhamento(
        df_sin_fechamento: pd.DataFrame,
        ws,
        nome_cc: str,
        pct: float,
        contas_a_receber: pd.DataFrame,
) -> None:
    """
    Escreve a aba Acompanhamento com estilo clean (preto/branco/cinza).

    Ordem de operações
    ──────────────────
    1. _construir_pivot_mensal → (df_pivot, col_map)
       Colunas internas sem '/' ("jan_25_liquido") — itertuples-safe.
       col_map guarda o label de display ("jan/25_liquido") para o cabeçalho.
    2. Soma *_liquido → valor_liquido_total; calcula repasse e pct.
    3. Merge CR pelo titulo — 1:1, sem inflação.
    4. Calcula contrato_total e check.
    5. Monta cabeçalho duplo (grupo de mês + métrica) usando col_map para exibição.
    6. Itera com itertuples usando os nomes internos (sem '/').
    """

    # ── 1. Pivot ──────────────────────────────────────────────────────────────

    df_sin_fechamento = df_sin_fechamento.rename(columns={
        "amortizacao": "bruto"
    })

    df_pivot, col_map = _construir_pivot_mensal(df_sin_fechamento)

    if df_pivot.empty:
        ws.cell(row=1, column=1, value="Sem dados para o acompanhamento.").font = \
            Font(italic=True, color="767676", size=10)
        return

    # ── 2. Totais históricos ──────────────────────────────────────────────────
    cols_liq = [c for c in df_pivot.columns if c.endswith("_liquido")]
    df_pivot["valor_liquido_total"] = df_pivot[cols_liq].sum(axis=1)
    df_pivot["pct_repasse"] = pct
    df_pivot["valor_liquido_repasse"] = (df_pivot["valor_liquido_total"] * pct).round(2)

    # ── 3. Merge CR (1 linha por titulo — sem inflação) ───────────────────────
    cr = (
        contas_a_receber
        .drop(columns=["cliente"], errors="ignore")
        [["titulo", "inadimplente", "a_vencer", "carteira_total", "vgv_vendido"]]
        .drop_duplicates(subset=["titulo"])
    )
    df_pivot = df_pivot.merge(cr, on="titulo", how="left")

    df_pivot = df_pivot.dropna(subset=['vgv_vendido']).reset_index(drop=True) # isso faz com que os clientes com contratos com distrato saiam

    # ── 4. Métricas derivadas ─────────────────────────────────────────────────
    df_pivot["contrato_total"] = (
            df_pivot["valor_liquido_total"] + df_pivot["carteira_total"].fillna(0)
    )
    df_pivot["check"] = (
            df_pivot["contrato_total"] - df_pivot["vgv_vendido"].fillna(0)
    )

    df_pivot["total_por_cliente"] = (df_pivot["vgv_vendido"] * pct).round(2)

    df_pivot['valor_a_repassar'] = (df_pivot["total_por_cliente"] - df_pivot['valor_liquido_repasse'])

    df_pivot = df_pivot.rename(columns={
        "amortizacao": "bruto"
    })

    # ── 5. Layout de colunas ──────────────────────────────────────────────────
    antes_validas = [c for c in COLS_ACOMPANHAMENTO_ANTES if c in df_pivot.columns]
    depois_validas = [c for c in COLS_ACOMPANHAMENTO_DEPOIS if c in df_pivot.columns]
    pivot_cols = [
        c for c in df_pivot.columns
        if c not in antes_validas and c not in depois_validas
    ]

    # Agrupa pivot_cols por mês usando col_map para obter o label "jan/25"
    # col_interna = "jan_25_liquido"  →  col_display = "jan/25_liquido"
    # label do grupo = "jan/25" (tudo antes do último '_')
    grupos_mes: dict[str, list[str]] = {}
    for col in pivot_cols:
        display = col_map.get(col, col)  # "jan/25_liquido"
        grupo = display.rsplit("_", 1)[0]  # "jan/25"
        grupos_mes.setdefault(grupo, []).append(col)

    todas_colunas = antes_validas + pivot_cols + depois_validas
    n = len(todas_colunas)

    # ── Linha 1: Título ───────────────────────────────────────────────────────
    _titulo_ws_cl(ws, f"Acompanhamento Mensal – {nome_cc}", n, 1)

    # ── Linha 2: Metadados ────────────────────────────────────────────────────
    n_clientes = df_pivot["cliente"].nunique() if "cliente" in df_pivot.columns else "–"
    _meta_ws_cl(ws, (
        f"% de repasse: {pct:.2%}   |   "
        f"Clientes: {n_clientes}   |   "
        f"Meses: {len(grupos_mes)}   |   "
        f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    ), n, 2)

    # ── Linhas 3–4: Cabeçalhos duplos ────────────────────────────────────────
    ROW_GRP = 3  # rótulo do mês ("JAN/25")
    ROW_HDR = 4  # sub-cabeçalho de métrica ("Liquido")
    col_idx = 1

    # Colunas fixas ANTES: mesclam 3+4 verticalmente
    for col in antes_validas:
        cab = col.replace("_", " ").title()
        ws.merge_cells(
            start_row=ROW_GRP, start_column=col_idx,
            end_row=ROW_HDR, end_column=col_idx,
        )
        _cabecalho_cl(ws, col_idx, ROW_GRP, cab, _HDR_FILL_CL)
        col_idx += 1

    # Grupos de mês: mescla horizontal linha 3; sub-cabeçalhos linha 4
    for grupo_label, cols_do_mes in grupos_mes.items():
        qtd = len(cols_do_mes)
        if qtd > 1:
            ws.merge_cells(
                start_row=ROW_GRP, start_column=col_idx,
                end_row=ROW_GRP, end_column=col_idx + qtd - 1,
            )
        _cabecalho_cl(ws, col_idx, ROW_GRP, grupo_label.upper(), _PIVOT_HDR_CL)
        for sub_col in cols_do_mes:
            # metrica = parte após o último '_' do display ("liquido")
            display = col_map.get(sub_col, sub_col)
            metrica = display.rsplit("_", 1)[-1].capitalize()
            _cabecalho_cl(ws, col_idx, ROW_HDR, metrica, _PIVOT_HDR_CL)
            col_idx += 1

    # Colunas fixas DEPOIS: mesclam 3+4 verticalmente
    for col in depois_validas:
        cab = col.replace("_", " ").title()
        ws.merge_cells(
            start_row=ROW_GRP, start_column=col_idx,
            end_row=ROW_HDR, end_column=col_idx,
        )
        _cabecalho_cl(ws, col_idx, ROW_GRP, cab, _HDR_FILL_CL)
        col_idx += 1

    ws.row_dimensions[ROW_GRP].height = 20
    ws.row_dimensions[ROW_HDR].height = 20

    # ── Dados ─────────────────────────────────────────────────────────────────
    colunas_valor_set = set(pivot_cols + depois_validas + ['vgv_vendido', 'total_por_cliente'])
    ri = ROW_HDR + 1

    for idx_linha, row in enumerate(df_pivot.itertuples(index=False)):
        fill_fundo = _PIVOT_ALT_CL if idx_linha % 2 == 0 else PatternFill()

        for ci, col in enumerate(todas_colunas, 1):
            val = getattr(row, col, None)  # col_interna sem '/' → funciona
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _BORDA_CL
            cell.font = Font(size=10, color="1A1A1A")
            cell.fill = fill_fundo

            if col in colunas_valor_set:
                if not val:
                    cell.font = Font(size=10, color="C0C0C0")
                cell.number_format = "0.00%" if col == "pct_repasse" else "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col == "cliente":
                cell.alignment = Alignment(horizontal="left")
            elif col == "novo_cliente":
                cell.alignment = Alignment(horizontal="center")
                if val == "S":
                    cell.font = Font(size=10, bold=True, color="404040")
            else:
                cell.alignment = Alignment(horizontal="center")

        ri += 1

    # ── Linha de total ────────────────────────────────────────────────────────
    _linha_total_cl(ws, ri, n, "TOTAL", df_pivot, todas_colunas, colunas_valor_set)

    # ── Larguras ──────────────────────────────────────────────────────────────
    larguras_fixas = {
        "cliente": 38, "titulo": 12, "novo_cliente": 8,
        "valor_liquido_total": 18, "valor_liquido_repasse": 18,
        "contrato_total": 16, "check": 14, "vgv_vendido": 16,
        "inadimplente": 16, "a_vencer": 14, "carteira_total": 16,
        "pct_repasse": 10,
    }
    for i, col in enumerate(todas_colunas, 1):
        ws.column_dimensions[get_column_letter(i)].width = (
            larguras_fixas.get(col, 14 if col in colunas_valor_set else 12)
        )

    ws.freeze_panes = f"A{ROW_HDR + 1}"


# ═════════════════════════════════════════════════════════════════════════════
# FECHAMENTO
# ═════════════════════════════════════════════════════════════════════════════

def _adicionar_fechamento_sintetico(wb: Workbook, df: pd.DataFrame,
                                    nome_cc: str, pct: float, centro: dict) -> None:
    ini, fim, mes_label = _mes_anterior()

    if "Fechamento" in wb.sheetnames:
        del wb["Fechamento"]
    ws = wb.create_sheet("Fechamento")

    cabecalhos, colunas_df, colunas_valor = _colunas_excel_sintetico(centro)

    cols_remover = {"inadimplente", "a_vencer", "carteira_total", "vgv_vendido", "contrato_total", "check"}
    colunas_df = [c for c in colunas_df if c not in cols_remover]
    colunas_valor = [c for c in colunas_valor if c not in cols_remover]
    cabecalhos = [h for h, c in zip(cabecalhos, _colunas_excel_sintetico(centro)[1])
                  if c not in cols_remover]

    n = len(cabecalhos)

    df_mes = df[df["dt_baixa"].dt.date.between(ini, fim)].copy()

    novos_nm = set(df[df["novo_cliente"] == "S"]["cliente"].unique())

    df_exibir = pd.concat([
        df[df["cliente"].isin(novos_nm)],
        df_mes[~df_mes["cliente"].isin(novos_nm)],
    ], ignore_index=True).sort_values(["cliente", "dt_baixa"]).reset_index(drop=True)

    cols_soma = [c for c in colunas_valor if c in df_exibir.columns]
    df_exibir = (
        df_exibir
        .groupby(["cliente", "novo_cliente"], as_index=False)
        .agg({**{c: "sum" for c in cols_soma}, "pct_repasse": "first"})
        .sort_values("cliente")
        .reset_index(drop=True)
    )

    ri = 1
    _titulo_ws(ws, f"Fechamento {mes_label}  –  {nome_cc}", n, ri);
    ri += 1
    _secao_ws(ws, f"▶  REPASSE DO MÊS ANTERIOR ({mes_label})  –  Todos os Clientes", n, ri);
    ri += 1
    _cabecalhos_ws(ws, cabecalhos, ri);
    ri += 1

    if len(df_exibir):
        for idx, row in enumerate(df_exibir.itertuples(index=False)):
            e_novo = getattr(row, "novo_cliente") == "S"
            fill = _NOVO_FILL if e_novo else (_ALT_FILL if idx % 2 == 0 else PatternFill())
            for ci, col in enumerate(colunas_df, 1):
                val = getattr(row, col)
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                _aplicar_celula(cell, col, colunas_valor, set(), e_novo)
            ri += 1
        _linha_total(ws, ri, n, f"TOTAL GERAL – {mes_label}", df_exibir,
                     colunas_df, colunas_valor, pct)
    else:
        ws.merge_cells(f"A{ri}:{get_column_letter(n)}{ri}")
        ws.cell(row=ri, column=1,
                value="Nenhum registro encontrado para o mês anterior.").font = \
            Font(italic=True, color="595959")

    larguras = {"cliente": 42, "novo_cliente": 8, "pct_repasse": 12, "valor_liquido_repasse": 18}
    _, _, colunas_num = _layout(centro)
    for c in colunas_num:
        larguras[c] = 14
    for i, col in enumerate(colunas_df, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 12)
    ws.freeze_panes = "A4"


def _adicionar_fechamento_analitico(wb: Workbook, df: pd.DataFrame,
                                    nome_cc: str, pct: float,
                                    centro: dict) -> tuple[float, float]:
    ini, fim, mes_label = _mes_anterior()

    if "Fechamento" in wb.sheetnames:
        del wb["Fechamento"]
    ws = wb.create_sheet("Fechamento")

    cabecalhos, colunas_df, colunas_valor = _colunas_excel_analitico(centro)
    colunas_data = {"dt_baixa", "data_vencimento"}
    n = len(cabecalhos)

    df_mes = df[df["dt_baixa"].dt.date.between(ini, fim)].copy()
    novos_nm = set(df[df["novo_cliente"] == "S"]["cliente"].unique())

    df_exibir = pd.concat([
        df[df["cliente"].isin(novos_nm)],
        df_mes[~df_mes["cliente"].isin(novos_nm)],
    ], ignore_index=True).sort_values(["cliente", "dt_baixa"])

    ri = 1
    _titulo_ws(ws, f"Fechamento {mes_label}  –  {nome_cc}", n, ri);
    ri += 1
    _secao_ws(ws, f"▶  REPASSE DO MÊS ANTERIOR ({mes_label})  –  Todos os Clientes", n, ri);
    ri += 1
    _cabecalhos_ws(ws, cabecalhos, ri);
    ri += 1

    if len(df_exibir):
        for cliente, grupo in df_exibir.groupby("cliente", sort=True):
            e_novo = (grupo["novo_cliente"] == "S").any()
            for idx, row in enumerate(grupo.itertuples(index=False)):
                fill = _NOVO_FILL if e_novo else (_ALT_FILL if idx % 2 == 0 else PatternFill())
                for ci, col in enumerate(colunas_df, 1):
                    val = getattr(row, col) if hasattr(row, col) else None
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill = fill
                    _aplicar_celula(cell, col, colunas_valor, colunas_data, e_novo)
                ri += 1
            sf = Font(bold=True, size=10, color="404040")
            for ci in range(1, n + 1):
                c = ws.cell(row=ri, column=ci)
                c.fill = _SUBTOT_FILL
                c.font = sf
                c.border = _BORDA
            ws.cell(row=ri, column=2, value=f"Subtotal – {cliente}").alignment = \
                Alignment(horizontal="left")
            for ci, col in enumerate(colunas_df, 1):
                if col in colunas_valor and col in grupo.columns:
                    c = ws.cell(row=ri, column=ci, value=grupo[col].sum())
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
                elif col == "pct_repasse":
                    c = ws.cell(row=ri, column=ci, value=pct)
                    c.number_format = "0.00%"
                    c.alignment = Alignment(horizontal="center")
            ri += 1

        _linha_total(ws, ri, n, f"TOTAL GERAL – {mes_label}", df_exibir,
                     colunas_df, colunas_valor, pct)
    else:
        ws.merge_cells(f"A{ri}:{get_column_letter(n)}{ri}")
        ws.cell(row=ri, column=1,
                value="Nenhum registro encontrado para o mês anterior.").font = \
            Font(italic=True, color="595959")

    larguras = {"dt_baixa": 12, "cliente": 40, "novo_cliente": 7,
                "documento": 16, "titulo": 14, "parcela": 10, "tc": 8,
                "unidade_principal": 14, "portador": 12, "operacao": 12,
                "data_vencimento": 12, "pct_repasse": 12, "valor_liquido_repasse": 18}
    _, _, colunas_num = _layout(centro)
    for c in colunas_num:
        larguras[c] = 14
    for i, col in enumerate(colunas_df, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 12)
    ws.freeze_panes = "A4"

    return (df_exibir["liquido"].sum(), df_exibir["valor_liquido_repasse"].sum()) \
        if len(df_exibir) else (0.0, 0.0)


# ═════════════════════════════════════════════════════════════════════════════
# ORQUESTRADOR PRINCIPAL
# ═════════════════════════════════════════════════════════════════════════════

def _obter_competencia():
    """
    Retorna informações padronizadas da competência atual.
    """

    # Linux / Mac
    try:
        locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")

    # Windows
    except:
        try:
            locale.setlocale(locale.LC_TIME, "Portuguese_Brazil.1252")
        except:
            pass

    hoje = datetime.now()

    competencia_dt = hoje - relativedelta(months=1)

    return {
        "ano": str(competencia_dt.year),
        "mes_ref": competencia_dt.strftime("%m. %B").capitalize(),
        "competencia": competencia_dt.strftime("%m.%Y"),
    }

def transformar_centro(centro: dict, novos: set[str], contas_a_receber, comp: dict) -> dict:

    nome_cc = str(centro["centro_custo"]).strip()
    slug_cc = _slug(nome_cc)
    nome_pasta_cc = slug_cc.replace("_", " ").title()
    pct = float(centro.get("pct_repasse") or 0)
    _, _, colunas_num = _layout(centro)

    dir_brutos = BASE_INPUT_DIR / slug_cc / "dados_brutos"
    dir_consol = BASE_OUTPUT_DIR / nome_pasta_cc / comp["ano"] / comp["mes_ref"]

    dir_consol.mkdir(parents=True, exist_ok=True)

    destino_ana = dir_consol / f"Analítico - {comp['competencia']}.xlsx"
    destino_sin = dir_consol / f"Sintético - {comp['competencia']}.xlsx"
    destino_acom = dir_consol / f"Acompanhamento - {comp['competencia']}.xlsx"

    resultado = {
        "nome_cc": nome_cc,
        "total_liquido_mes_ana": 0.0,
        "total_mes_repasse_ana": 0.0,
        "pct_repasse": pct,
        "df_ana": pd.DataFrame(),
    }

    # ── 1. Lê todos os analíticos brutos ──────────────────────────────────────
    arqs_ana = sorted(dir_brutos.glob(f"{slug_cc}_*_analitico.xlsx"))
    if not arqs_ana:
        logger.warning("[%s] Nenhum analítico encontrado", nome_cc)
        return resultado

    frames_ana = []
    for arq in arqs_ana:
        try:
            frames_ana.append(ler_analitico(arq, novos, centro))
            logger.info("  ✓ %s", arq.name)
        except Exception:
            logger.exception("  ✗ %s — pulando", arq.name)

    if not frames_ana:
        return resultado

    df_ana = pd.concat(frames_ana, ignore_index=True)
    df_ana["pct_repasse"] = pct
    df_ana["valor_liquido_repasse"] = (df_ana["liquido"] * pct).round(2)
    df_ana = df_ana.sort_values(["cliente", "dt_baixa"]).reset_index(drop=True)

    # ── 2. Deriva o sintético temporal (base do fechamento e acompanhamento) ──

    cols_agg = {c: "sum" for c in colunas_num if c in df_ana.columns}

    df_sin_fechamento = (
        df_ana
        .groupby(["dt_baixa", "cliente", "novo_cliente", "titulo"], as_index=False)
        .agg(cols_agg)
    )

    if "liquido" not in df_sin_fechamento.columns and "vl_baixa" in df_sin_fechamento.columns:
        df_sin_fechamento["liquido"] = df_sin_fechamento["vl_baixa"]

    df_sin_fechamento["pct_repasse"] = pct
    df_sin_fechamento["valor_liquido_repasse"] = (
            df_sin_fechamento["liquido"] * pct
    ).round(2)
    df_sin_fechamento = df_sin_fechamento.sort_values("cliente").reset_index(drop=True)

    # ── 3. Sintético tradicional (1 linha por cliente) ────────────────────────

    df_sin = (
        df_sin_fechamento
        .groupby(["cliente", "novo_cliente", "titulo"], as_index=False)
        .agg({
            **{c: "sum" for c in colunas_num if c in df_sin_fechamento.columns},
            "valor_liquido_repasse": "sum",
            "pct_repasse": "first",
        })
        .sort_values("cliente")
        .reset_index(drop=True)
    )

    def remover_aba(wb) -> None:
        """
        Deixa somenete em a aba de fechamento nas planilhas, exceto naquelas que estão dentro da lista de verificação.
        :param wb:
        :return:  None
        """
        for wb_aba in wb.sheetnames:
            if not nome_cc in ["GRAND PALADIUM", "LAGOON CLUB RESIDENCE"] and wb_aba != "Fechamento":
                logger.info(f"Removendo aba do centro → %s", nome_cc)

                wb.remove(wb[wb_aba])

    # ── 4. Salva analítico (com fechamento) ───────────────────────────────────
    wb_ana = Workbook()
    ws_ana = wb_ana.active
    ws_ana.title = "Analítico"
    _salvar_analitico(df_ana, ws_ana, nome_cc, pct, centro)
    total_mes, total_mes_repasse = _adicionar_fechamento_analitico(
        wb_ana, df_ana, nome_cc, pct, centro
    )

    remover_aba(wb_ana)
    wb_ana.save(destino_ana)

    logger.info("[%s] Analítico salvo → %s", nome_cc, destino_ana.name)

    resultado["total_liquido_mes_ana"] = total_mes
    resultado["total_mes_repasse_ana"] = total_mes_repasse
    resultado["df_ana"] = df_ana

    # ── 5. Salva sintético derivado (com fechamento) ──────────────────────────
    wb_sin = Workbook()
    ws_sin = wb_sin.active
    ws_sin.title = "Consolidado"
    _salvar_sintetico(df_sin, ws_sin, nome_cc, pct, centro)
    _adicionar_fechamento_sintetico(wb_sin, df_sin_fechamento, nome_cc, pct, centro)

    remover_aba(wb_sin)
    wb_sin.save(destino_sin)

    logger.info("[%s] Sintético salvo → %s", nome_cc, destino_sin.name)

    # ── 6. Salva acompanhamento mensal (pivot clean) ──────────────────────────
    wb_acom = Workbook()
    ws_acom = wb_acom.active
    ws_acom.title = "Acompanhamento"
    _salvar_acompanhamento(df_sin_fechamento, ws_acom, nome_cc, pct, contas_a_receber=contas_a_receber)
    wb_acom.save(destino_acom)
    logger.info("[%s] Acompanhamento salvo → %s", nome_cc, destino_acom.name)

    return resultado


# ═════════════════════════════════════════════════════════════════════════════
# CONSOLIDADO GERAL ANALÍTICO  (para Power BI)
# ═════════════════════════════════════════════════════════════════════════════

def _gerar_consolidado_geral(resultados: list[dict]) -> None:
    frames = [
        r["df_ana"].assign(centro_custo=r["nome_cc"])
        for r in resultados
        if isinstance(r.get("df_ana"), pd.DataFrame) and len(r["df_ana"])
    ]

    if not frames:
        logger.warning("Nenhum analítico disponível para o consolidado geral")
        return

    df = pd.concat(frames, ignore_index=True)

    cols_ordem = ["centro_custo", "periodo", "dt_baixa", "cliente", "novo_cliente"]
    cols_resto = [c for c in df.columns if c not in cols_ordem]
    df = df[[c for c in cols_ordem if c in df.columns] + cols_resto]
    df = df.sort_values(
        [c for c in ["centro_custo", "cliente", "dt_baixa"] if c in df.columns]
    ).reset_index(drop=True)

    df["data_carga"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    destino = BASE_OUTPUT_DIR / "consolidado_geral_analitico.csv"
    df.to_csv(destino, sep=";", decimal=",", index=False)
    logger.info("Consolidado geral → %s  (%d linhas, %d centros)",
                destino.name, len(df), len(frames))


# ═════════════════════════════════════════════════════════════════════════════
# RESUMO WHATSAPP
# ═════════════════════════════════════════════════════════════════════════════
def _fmt_moeda(valor: float) -> str:
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _imprimir_resumo(resultados: list[dict]) -> None:
    _, _, mes_label = _mes_anterior()

    print("\n" + "=" * 60)
    print(f"RESUMO – {mes_label}")
    print("=" * 60)

    for r in resultados:
        total_liquido = r.get("total_liquido_mes_ana", 0.0)
        total_mes_repasse_ana = r.get("total_mes_repasse_ana", 0.0)
        pct_repasse = r.get("pct_repasse", 0.0)

        print(f"\n{r['nome_cc']}")
        print(f"Competência: {mes_label}")
        print(f"* Total líquido do mês fechado: R$ {_fmt_moeda(total_liquido)}")
        print(f"* Percentual de repasse: {pct_repasse:.2%}")
        print(f"* Valor líquido de repasse: R$ {_fmt_moeda(total_mes_repasse_ana)}")
        print("—" * 40)

    print("=" * 60 + "\n")


# ═════════════════════════════════════════════════════════════════════════════
# ENTRYPOINT
# ═════════════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(description="Transform – Contas Recebidas")
    parser.add_argument("--so-consolidar", action="store_true",
                        help="Apenas consolida os brutos (sem fechamento nem painel)")
    args = parser.parse_args()

    centros = carregar_centros_ativos()
    logger.info("%d centros de custo ativos", len(centros))

    resultados = []
    contas_a_receber = ler_contas_a_receber()

    comp = _obter_competencia()
    for centro in centros:
        nome_cc = str(centro["centro_custo"]).strip()
        try:
            novos = carregar_extrato(nome_cc)
            res = transformar_centro(centro, novos, contas_a_receber, comp)
            resultados.append(res)
        except Exception:
            logger.exception("Erro em '%s'", nome_cc)

    if not args.so_consolidar:
        _gerar_consolidado_geral(resultados)
        _imprimir_resumo(resultados)

    logger.info("Transform concluído.")


if __name__ == "__main__":
    main()

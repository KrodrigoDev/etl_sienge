"""
stages/transform/transform_contas_recebidas.py
------------------------------------------------
Transform – Contas Recebidas

Para cada centro de custo ativo:
  1. Carrega o extrato de empreendimentos (aba UNIDADES) para identificar
     clientes com CRI no mês anterior ao vigente → flag novo_cliente
  2. Lê todos os sintéticos / analíticos de dados_brutos/
  3. Faz join pelo nome do mutuário e adiciona coluna novo_cliente (S/N)
  4. Aplica % de repasse → valor_liquido_repasse
  5. Salva consolidado sintético e analítico em dados_consolidados/

Regra novo_cliente
──────────────────
  • Extrai o nome do campo "cliente" do Sienge removendo " (código)" do final
  • Compara com Nome Mutuário do extrato (normalizado: upper + strip)
  • Se a Data de Inclusão dos Dados de Registro (CRI) cair no mês anterior
    ao vigente → novo_cliente = "S", caso contrário "N"
  • O extrato também é filtrado pelo empreendimento para evitar falsos positivos

Linha de início dos dados
──────────────────────────
  • Com empresa preenchida nos filtros → dados a partir do índice 10
  • Sem empresa (relatório agrega todas) → dados a partir do índice 9
  O campo 'possui_empresa' do auxiliar controla isso (padrão: True)
"""

from __future__ import annotations

import logging
import re
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

AUXILIAR_PATH = (
        Path(__file__).resolve().parents[2]
        / "stages" / "rpa" / "files" / "reference" / "auxiliar.xlsx"
)
BASE_OUTPUT_DIR = (
        Path(__file__).resolve().parents[2]
        / "stages" / "rpa" / "files" / "output" / "contas_recebidas"
)

EXTRATO_PATH = (
        Path(__file__).resolve().parents[2]
        / "stages" / "rpa" / "files" / "reference" / "extrato_empreendimentos.xls"
)

RODAPE_PREFIXOS = (
    "Total geral", "Total da empresa", "Total de parcelas",
    "Total de títulos", "(*)", "(C)", "(S)", "SIENGE",
)
_RE_TIMESTAMP = re.compile(r"^\d{2}/\d{2}/\d{4} - \d{2}:\d{2}:\d{2}$")
_RE_CODIGO = re.compile(r"\s*\(\d+\)\s*$")  # remove " (12345)" do fim

# ── Estilos compartilhados ────────────────────────────────────────────────────
_THIN = Side(style="thin", color="BFBFBF")
_BORDA = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_FILL = PatternFill("solid", start_color="2E75B6")
_TITLE_FILL = PatternFill("solid", start_color="1F4E79")
_ALT_FILL = PatternFill("solid", start_color="DCE6F1")
_TOTAL_FILL = PatternFill("solid", start_color="1F4E79")
_SUBTOT_FILL = PatternFill("solid", start_color="BDD7EE")  # azul claro para subtotais
_NOVO_FILL = PatternFill("solid", start_color="FFF2CC")  # amarelo para novos clientes


# ── Helpers ───────────────────────────────────────────────────────────────────

def _slug(nome: str) -> str:
    return nome.strip().lower().replace(" ", "_").replace("-", "")[:40]


def _nome_sienge(cliente_str: str) -> str:
    """Remove o código entre parênteses do final: 'JOSE SILVA (123)' → 'JOSE SILVA'."""
    return _RE_CODIGO.sub("", str(cliente_str)).strip().upper()


def _mes_anterior() -> tuple[date, date]:
    hoje = date.today()
    ini = (hoje.replace(day=1) - timedelta(days=1)).replace(day=1)
    fim = ini + relativedelta(months=1) - timedelta(days=1)
    return ini, fim


def _eh_rodape(valor) -> bool:
    v = str(valor).strip()
    return v.startswith(RODAPE_PREFIXOS) or bool(_RE_TIMESTAMP.match(v))


def _inicio_dados(caminho: Path) -> int:
    """
    Detecta automaticamente a linha onde começam os dados.
    Com empresa: linha 10 (índice 10).
    Sem empresa: linha 9 (índice 9).
    Critério: se a célula A10 for NaN ou rodapé, usa 9.
    """
    df_raw = pd.read_excel(caminho, sheet_name="Relatório", header=None, nrows=12)
    val_10 = str(df_raw.iloc[10, 0]).strip() if len(df_raw) > 10 else ""
    if not val_10 or val_10 == "nan" or _eh_rodape(val_10):
        return 9
    return 10


# ── Extrato de empreendimentos ────────────────────────────────────────────────

def carregar_extrato(nome_empreendimento: str) -> set[str]:
    """
    Retorna conjunto de nomes de mutuários (upper) cujo CRI caiu
    no mês anterior ao vigente, filtrado pelo empreendimento.
    """
    if not EXTRATO_PATH.exists():
        logger.warning("Extrato não encontrado em %s — novo_cliente sempre 'N'", EXTRATO_PATH)
        return set()

    df = pd.read_excel(EXTRATO_PATH, sheet_name="UNIDADES", header=1)
    df.columns = ["cnpj", "empreendimento", "contrato",
                  "nome_mutuario", "cpf", "dt_assinatura", "dt_cri"]
    df["dt_cri"] = pd.to_datetime(df["dt_cri"], errors="coerce")
    df["nome_up"] = df["nome_mutuario"].astype(str).str.strip().str.upper()

    # Filtra pelo empreendimento (match parcial, case-insensitive)
    nome_up = nome_empreendimento.upper()
    df = df[df["empreendimento"].astype(str).str.upper().str.contains(
        re.escape(nome_up[:15]), regex=True  # usa os 15 primeiros chars como chave
    )]

    ini, fim = _mes_anterior()
    mask = (df["dt_cri"].dt.date >= ini) & (df["dt_cri"].dt.date <= fim)
    novos = set(df.loc[mask, "nome_up"].tolist())
    logger.info("Extrato [%s]: %d cliente(s) novo(s) em %s/%s",
                nome_empreendimento, len(novos), ini.strftime("%m"), ini.strftime("%Y"))
    return novos


# ── Leitura do auxiliar ───────────────────────────────────────────────────────

def carregar_centros_ativos() -> list[dict]:
    df = pd.read_excel(AUXILIAR_PATH, sheet_name="centros_custo")
    df = df[df["ativo"].str.strip().str.lower() == "sim"]
    return df.to_dict(orient="records")


# ── Mapeamentos de colunas ────────────────────────────────────────────────────
# Verificado contra os arquivos reais (row[9] = linha de cabeçalho do Sienge)

# Sintético – Juros embutidos
# col[0]=Cliente col[5]=Amortização col[7]=Juros col[8]=Correção
# col[9]=Acréscimo col[11]=Seguro col[12]=Taxa adm col[13]=Desconto col[14]=Líquido
COLUNAS_SINTETICO_JUROS = {
    0: "cliente", 5: "amortizacao", 7: "juros", 8: "correcao",
    9: "acrescimo", 11: "seguro", 12: "taxa_adm", 13: "desconto", 14: "liquido",
}

# Sintético – Padrão
# col[0]=Cliente col[5]=Vl.baixa col[7]=Acréscimo col[8]=Seguro
# col[10]=Taxa adm col[11]=Desconto col[12]=Líquido
COLUNAS_SINTETICO_PADRAO = {
    0: "cliente", 5: "vl_baixa", 7: "acrescimo", 8: "seguro",
    10: "taxa_adm", 11: "desconto", 12: "liquido",
}

# Analítico – Juros embutidos
# col[0]=Dt.baixa col[1]=Cliente col[4]=Documento col[7]=Título col[8]=Parc
# col[9]=TC col[10]=Unid.princ col[12]=Port col[13]=Oper col[14]=Data vecto
# col[15]=Amortização col[16]=Juros contr col[18]=Correção col[19]=Acréscimos
# col[20]=Seguro col[21]=Taxa adm col[22]=Desconto col[23]=Líquido
COLUNAS_ANALITICO_JUROS = {
    0: "dt_baixa", 1: "cliente", 4: "documento", 7: "titulo",
    8: "parcela", 9: "tc", 10: "unidade_principal", 12: "portador",
    13: "operacao", 14: "data_vencimento", 15: "amortizacao", 16: "juros",
    18: "correcao", 19: "acrescimo", 20: "seguro", 21: "taxa_adm",
    22: "desconto", 23: "liquido",
}

# Analítico – Padrão
# col[0]=Dt.baixa col[1]=Cliente col[4]=Dt.emissão col[5]=Documento col[7]=Título
# col[8]=Parc col[9]=TC col[10]=Unid.princ col[11]=Port col[12]=Oper
# col[13]=Data vecto col[14]=Vl.baixa col[15]=Acréscimo col[16]=Seguro
# col[17]=Taxa adm col[18]=Desconto col[19]=Líquido
COLUNAS_ANALITICO_PADRAO = {
    0: "dt_baixa", 1: "cliente", 4: "dt_emissao", 5: "documento",
    7: "titulo", 8: "parcela", 9: "tc", 10: "unidade_principal",
    11: "portador", 12: "operacao", 13: "data_vencimento", 14: "vl_baixa",
    15: "acrescimo", 16: "seguro", 17: "taxa_adm", 18: "desconto", 19: "liquido",
}

# Colunas numéricas comuns a ambos os layouts
COLUNAS_NUM_JUROS = ["amortizacao", "juros", "correcao", "acrescimo",
                     "seguro", "taxa_adm", "desconto", "liquido"]
COLUNAS_NUM_PADRAO = ["vl_baixa", "acrescimo", "seguro",
                      "taxa_adm", "desconto", "liquido"]

# Mapa: tipo_coluna do auxiliar → mapeamentos (sintetico, analitico, colunas_num)
LAYOUT = {
    "juros embutidos": (COLUNAS_SINTETICO_JUROS, COLUNAS_ANALITICO_JUROS, COLUNAS_NUM_JUROS),
    "padrão": (COLUNAS_SINTETICO_PADRAO, COLUNAS_ANALITICO_PADRAO, COLUNAS_NUM_PADRAO),
}


def _layout(centro: dict) -> tuple[dict, dict, list[str]]:
    """Retorna (cols_sin, cols_ana, colunas_num) conforme tipo_coluna do auxiliar."""
    tipo = str(centro.get("tipo_coluna", "")).strip().lower()
    return LAYOUT.get(tipo, LAYOUT["juros embutidos"])


def extrair_periodo(df_raw: pd.DataFrame) -> str:
    try:
        texto = str(df_raw.iloc[6, 2])
        m = re.search(r"(\d{2})/(\d{2})/(\d{4})", texto)
        if m:
            return f"{m.group(3)}{m.group(2)}"
    except Exception:
        pass
    return "000000"


def ler_sintetico(caminho: Path, novos: set[str], centro: dict) -> pd.DataFrame:
    cols_sin, _, colunas_num = _layout(centro)

    df_raw = pd.read_excel(caminho, sheet_name="Relatório", header=None)
    periodo = extrair_periodo(df_raw)
    inicio = _inicio_dados(caminho)

    dados = df_raw.iloc[inicio:][list(cols_sin.keys())].copy()
    dados.columns = list(cols_sin.values())

    # Descarta linhas de rodapé/totais (incluindo "Total do cliente")
    dados = dados[dados["cliente"].notna()]
    dados = dados[~dados["cliente"].astype(str).str.strip().eq("")]
    dados = dados[~dados["cliente"].astype(str).apply(_eh_rodape)]

    for col in colunas_num:
        dados[col] = pd.to_numeric(dados[col], errors="coerce").fillna(0.0)

    # Garante que "liquido" sempre existe (alias de vl_baixa no layout padrão)
    if "liquido" not in dados.columns and "vl_baixa" in dados.columns:
        dados["liquido"] = dados["vl_baixa"]

    dados.insert(0, "periodo", periodo)
    dados["novo_cliente"] = dados["cliente"].apply(
        lambda c: "S" if _nome_sienge(c) in novos else "N"
    )
    return dados.reset_index(drop=True)


def ler_analitico(caminho: Path, novos: set[str], centro: dict) -> pd.DataFrame:
    _, cols_ana, colunas_num = _layout(centro)

    df_raw = pd.read_excel(caminho, sheet_name="Relatório", header=None)
    periodo = extrair_periodo(df_raw)
    inicio = _inicio_dados(caminho)

    dados = df_raw.iloc[inicio:][list(cols_ana.keys())].copy()
    dados.columns = list(cols_ana.values())

    # Descarta linhas de rodapé/totais E linhas "Total do cliente" do Sienge
    dados = dados[dados["cliente"].notna()]
    dados = dados[~dados["cliente"].astype(str).str.strip().eq("")]
    dados = dados[~dados["cliente"].astype(str).str.startswith("Total do cliente")]
    dados = dados[~dados["cliente"].astype(str).apply(_eh_rodape)]

    for col in colunas_num:
        if col == "amortizacao":
            dados[col] = (
                dados[col].astype(str)
                .str.replace(" P", "", regex=False).str.strip()
            )
            mask = dados[col].str.contains(",", na=False)
            dados.loc[mask, col] = (
                dados.loc[mask, col]
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
        dados[col] = pd.to_numeric(dados[col], errors="coerce").fillna(0.0)

    # Garante que "liquido" sempre existe
    if "liquido" not in dados.columns and "vl_baixa" in dados.columns:
        dados["liquido"] = dados["vl_baixa"]

    dados["dt_baixa"] = pd.to_datetime(dados["dt_baixa"], format="%d/%m/%Y", errors="coerce")

    dados.insert(0, "periodo", periodo)
    dados["novo_cliente"] = dados["cliente"].apply(
        lambda c: "S" if _nome_sienge(c) in novos else "N"
    )
    return dados.reset_index(drop=True)


# ── Transform ─────────────────────────────────────────────────────────────────

def transformar_centro_sintetico(centro: dict, novos: set[str]) -> None:
    nome_cc = str(centro["centro_custo"]).strip()
    slug_cc = _slug(nome_cc)
    pct = float(centro.get("pct_repasse") or 0)

    _, _, colunas_num = _layout(centro)

    dir_brutos = BASE_OUTPUT_DIR / slug_cc / "dados_brutos"
    dir_consol = BASE_OUTPUT_DIR / slug_cc / "dados_consolidados"
    dir_consol.mkdir(parents=True, exist_ok=True)

    arquivos = sorted(dir_brutos.glob(f"{slug_cc}_*_sintetico.xlsx"))
    if not arquivos:
        logger.warning("[%s] Nenhum sintético encontrado", nome_cc)
        return

    frames = []
    for arq in arquivos:
        try:
            frames.append(ler_sintetico(arq, novos, centro))
            logger.info("  ✓ %s", arq.name)
        except Exception:
            logger.exception("  ✗ %s — pulando", arq.name)

    if not frames:
        return

    raw = pd.concat(frames, ignore_index=True)

    consolidado = (
        raw.groupby(["cliente", "novo_cliente"], as_index=False)
        .agg({c: "sum" for c in colunas_num})
    )
    consolidado["pct_repasse"] = pct
    consolidado["valor_liquido_repasse"] = (consolidado["liquido"] * pct).round(2)
    consolidado = consolidado.sort_values("cliente").reset_index(drop=True)

    destino = dir_consol / f"{slug_cc}_consolidado_sintetico.xlsx"
    _salvar_sintetico(consolidado, destino, nome_cc, pct)
    logger.info("[%s] Sintético salvo → %s", nome_cc, destino.name)


def transformar_centro_analitico(centro: dict, novos: set[str]) -> None:
    nome_cc = str(centro["centro_custo"]).strip()
    slug_cc = _slug(nome_cc)
    pct = float(centro.get("pct_repasse") or 0)

    dir_brutos = BASE_OUTPUT_DIR / slug_cc / "dados_brutos"
    dir_consol = BASE_OUTPUT_DIR / slug_cc / "dados_consolidados"
    dir_consol.mkdir(parents=True, exist_ok=True)

    arquivos = sorted(dir_brutos.glob(f"{slug_cc}_*_analitico.xlsx"))
    if not arquivos:
        logger.warning("[%s] Nenhum analítico encontrado", nome_cc)
        return

    frames = []
    for arq in arquivos:
        try:
            frames.append(ler_analitico(arq, novos, centro))
            logger.info("  ✓ %s", arq.name)
        except Exception:
            logger.exception("  ✗ %s — pulando", arq.name)

    if not frames:
        return

    consolidado = pd.concat(frames, ignore_index=True)
    consolidado["pct_repasse"] = pct
    consolidado["valor_liquido_repasse"] = (consolidado["liquido"] * pct).round(2)
    consolidado = consolidado.sort_values(["cliente", "dt_baixa"]).reset_index(drop=True)

    destino = dir_consol / f"{slug_cc}_consolidado_analitico.xlsx"
    _salvar_analitico(consolidado, destino, nome_cc, pct)
    logger.info("[%s] Analítico salvo → %s", nome_cc, destino.name)


# ── Geração Excel – Sintético ─────────────────────────────────────────────────

def _salvar_sintetico(df: pd.DataFrame, destino: Path, nome_cc: str, pct: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    cabecalhos = [
        "Cliente", "Novo?", "Amortização", "Juros", "Correção",
        "Acréscimo", "Seguro", "Taxa adm", "Desconto", "Líquido",
        "% Repasse", "Vlr. Líq. Repasse",
    ]
    colunas_df = [
        "cliente", "novo_cliente", "amortizacao", "juros", "correcao",
        "acrescimo", "seguro", "taxa_adm", "desconto", "liquido",
        "pct_repasse", "valor_liquido_repasse",
    ]
    n_cols = len(cabecalhos)
    span = f"A1:{get_column_letter(n_cols)}1"

    # Título
    ws.merge_cells(span)
    ws["A1"] = f"Contas Recebidas – {nome_cc}"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Metadados
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    ws["A2"] = (
        f"% de repasse: {pct:.2%}   |   "
        f"Clientes únicos: {df['cliente'].nunique()}   |   "
        f"Novos clientes (CRI mês ant.): {(df['novo_cliente'] == 'S').sum()}   |   "
        f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Cabeçalhos
    for ci, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=3, column=ci, value=cab)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDA
    ws.row_dimensions[3].height = 30

    colunas_valor = {"amortizacao", "juros", "correcao", "acrescimo",
                     "seguro", "taxa_adm", "desconto", "liquido", "valor_liquido_repasse"}

    # Dados
    for ri, row in enumerate(df.itertuples(index=False), 4):
        e_novo = getattr(row, "novo_cliente") == "S"
        fill = _NOVO_FILL if e_novo else (_ALT_FILL if ri % 2 == 0 else PatternFill())
        for ci, col in enumerate(colunas_df, 1):
            val = getattr(row, col)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _BORDA
            cell.fill = fill
            cell.font = Font(size=10)
            if col == "cliente":
                cell.alignment = Alignment(horizontal="left")
            elif col == "novo_cliente":
                cell.alignment = Alignment(horizontal="center")
                if e_novo:
                    cell.font = Font(size=10, bold=True, color="7F6000")
            elif col == "pct_repasse":
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="center")
            elif col in colunas_valor:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # Totais
    tr = len(df) + 4
    tot_font = Font(bold=True, color="FFFFFF", size=10)
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=tr, column=ci)
        c.fill = _TOTAL_FILL;
        c.font = tot_font;
        c.border = _BORDA

    ws.cell(row=tr, column=1, value="TOTAL").alignment = Alignment(horizontal="center")
    for ci, col in {3: "amortizacao", 4: "juros", 5: "correcao", 6: "acrescimo",
                    7: "seguro", 8: "taxa_adm", 9: "desconto", 10: "liquido",
                    12: "valor_liquido_repasse"}.items():
        c = ws.cell(row=tr, column=ci, value=df[col].sum())
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right")
    c_pct = ws.cell(row=tr, column=11, value=pct)
    c_pct.number_format = "0.00%"
    c_pct.alignment = Alignment(horizontal="center")

    # Larguras
    for i, w in enumerate([42, 8, 14, 10, 10, 12, 10, 10, 10, 14, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    wb.save(destino)


# ── Geração Excel – Analítico ─────────────────────────────────────────────────

def _salvar_analitico(df: pd.DataFrame, destino: Path, nome_cc: str, pct: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Analítico"

    cabecalhos = [
        "Dt. Baixa", "Cliente", "Novo?", "Documento", "Título", "Parcela", "TC",
        "Unidade", "Portador", "Operação", "Vencimento",
        "Amortização", "Juros", "Correção", "Acréscimo", "Seguro", "Taxa Adm",
        "Desconto", "Líquido", "% Repasse", "Vlr. Líq. Repasse",
    ]
    colunas_df = [
        "dt_baixa", "cliente", "novo_cliente", "documento", "titulo", "parcela", "tc",
        "unidade_principal", "portador", "operacao", "data_vencimento",
        "amortizacao", "juros", "correcao", "acrescimo", "seguro", "taxa_adm",
        "desconto", "liquido", "pct_repasse", "valor_liquido_repasse",
    ]
    n_cols = len(cabecalhos)

    # Título
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"] = f"Contas Recebidas Analítico – {nome_cc}"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Metadados
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    ws["A2"] = (
        f"% de repasse: {pct:.2%}   |   "
        f"Registros: {len(df)}   |   "
        f"Clientes únicos: {df['cliente'].nunique()}   |   "
        f"Novos clientes (CRI mês ant.): {(df['novo_cliente'] == 'S').sum()}   |   "
        f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Cabeçalhos
    for ci, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=3, column=ci, value=cab)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDA
    ws.row_dimensions[3].height = 30

    colunas_valor = {"amortizacao", "juros", "correcao", "acrescimo", "seguro",
                     "taxa_adm", "desconto", "liquido", "valor_liquido_repasse"}
    colunas_data = {"dt_baixa", "data_vencimento"}

    # Dados + subtotal por cliente
    ri = 4
    for cliente, grupo in df.groupby("cliente", sort=True):
        e_novo = (grupo["novo_cliente"] == "S").any()
        fill_base = _NOVO_FILL if e_novo else None  # None → alterna por linha

        for row in grupo.itertuples(index=False):
            fill = _NOVO_FILL if e_novo else (_ALT_FILL if ri % 2 == 0 else PatternFill())
            for ci, col in enumerate(colunas_df, 1):
                val = getattr(row, col)
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = _BORDA
                cell.fill = fill
                cell.font = Font(size=10)
                if col in colunas_data:
                    cell.number_format = "dd/mm/yyyy"
                    cell.alignment = Alignment(horizontal="center")
                elif col == "pct_repasse":
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="center")
                elif col in colunas_valor:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col == "cliente":
                    cell.alignment = Alignment(horizontal="left")
                elif col == "novo_cliente":
                    cell.alignment = Alignment(horizontal="center")
                    if e_novo:
                        cell.font = Font(size=10, bold=True, color="7F6000")
                else:
                    cell.alignment = Alignment(horizontal="center")
            ri += 1

        # ── Subtotal do cliente ───────────────────────────────────────────────
        subtot_font = Font(bold=True, size=10,
                           color="7F6000" if e_novo else "1F4E79")
        for ci in range(1, n_cols + 1):
            c = ws.cell(row=ri, column=ci)
            c.fill = _SUBTOT_FILL
            c.font = subtot_font
            c.border = _BORDA

        # Label do subtotal (coluna "Cliente")
        ws.cell(row=ri, column=2, value=f"Subtotal – {cliente}").alignment = \
            Alignment(horizontal="left")
        ws.cell(row=ri, column=3, value="S" if e_novo else "N").alignment = \
            Alignment(horizontal="center")

        soma_cols = {12: "amortizacao", 13: "juros", 14: "correcao", 15: "acrescimo",
                     16: "seguro", 17: "taxa_adm", 18: "desconto", 19: "liquido",
                     21: "valor_liquido_repasse"}
        for ci, col in soma_cols.items():
            c = ws.cell(row=ri, column=ci, value=grupo[col].sum())
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        ws.cell(row=ri, column=20, value=pct).number_format = "0.00%"
        ri += 1

    # ── Total geral ───────────────────────────────────────────────────────────
    tot_font = Font(bold=True, color="FFFFFF", size=10)
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=ri, column=ci)
        c.fill = _TOTAL_FILL;
        c.font = tot_font;
        c.border = _BORDA

    ws.cell(row=ri, column=1, value="TOTAL GERAL").alignment = \
        Alignment(horizontal="center")
    for ci, col in {12: "amortizacao", 13: "juros", 14: "correcao", 15: "acrescimo",
                    16: "seguro", 17: "taxa_adm", 18: "desconto", 19: "liquido",
                    21: "valor_liquido_repasse"}.items():
        c = ws.cell(row=ri, column=ci, value=df[col].sum())
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right")
    ws.cell(row=ri, column=20, value=pct).number_format = "0.00%"

    # Larguras
    for i, w in enumerate([12, 40, 7, 16, 14, 10, 8, 14, 12, 12, 12,
                           14, 12, 12, 12, 10, 12, 12, 14, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    wb.save(destino)


# ── Entrypoint ────────────────────────────────────────────────────────────────

def main() -> None:
    centros = carregar_centros_ativos()
    logger.info("%d centros de custo ativos", len(centros))

    for centro in centros:
        nome_cc = str(centro["centro_custo"]).strip()
        try:
            novos = carregar_extrato(nome_cc)
            transformar_centro_sintetico(centro, novos)
            transformar_centro_analitico(centro, novos)
        except Exception:
            logger.exception("Erro no transform de '%s'", nome_cc)

    logger.info("Transform concluído.")


if __name__ == "__main__":
    main()

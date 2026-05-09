"""
transform_contas_recebidas_fechamento.py
────────────────────────────────────────
Adiciona aba 'Fechamento' nos consolidados sintético e analítico.

Sintético → Seção 1: repasse acumulado de todos os clientes no mês anterior
            Seção 2: somatório acumulado dos novos clientes (CRI mês ant.)

Analítico → Seção 1: registros com dt_baixa no mês anterior (com subtotais)
            Seção 2: série histórica completa dos novos clientes (com subtotais)

Ao final imprime resumo formatado para colar no WhatsApp.

Uso:
    python transform_contas_recebidas_fechamento.py  [sintetico.xlsx] [analitico.xlsx]

    Se os caminhos não forem informados, assume os arquivos com o padrão
    *_consolidado_sintetico.xlsx e *_consolidado_analitico.xlsx no diretório
    corrente.
"""
from __future__ import annotations

import logging
import sys
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Estilos ──────────────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="BFBFBF")
_BORDA = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDA_MED = Border(
    left=Side(style="medium", color="1F4E79"),
    right=Side(style="medium", color="1F4E79"),
    top=Side(style="medium", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_FILL = PatternFill("solid", start_color="2E75B6")
_TITLE_FILL = PatternFill("solid", start_color="1F4E79")
_ALT_FILL = PatternFill("solid", start_color="DCE6F1")
_TOTAL_FILL = PatternFill("solid", start_color="1F4E79")
_SUBTOT_FILL = PatternFill("solid", start_color="BDD7EE")
_NOVO_FILL = PatternFill("solid", start_color="FFF2CC")
_SECT_FILL = PatternFill("solid", start_color="E2EFDA")
_NOVO_TOT_FILL = PatternFill("solid", start_color="FFD966")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _mes_anterior() -> tuple[str, "pd.Period"]:
    hoje = date.today()
    ini = (hoje.replace(day=1) - timedelta(days=1)).replace(day=1)
    return ini.strftime("%m/%Y"), pd.Period(ini.strftime("%Y-%m"), "M")


def _nome_cc_da_planilha(caminho: Path, sheet: str, prefixo: str) -> str:
    try:
        wb = load_workbook(caminho, read_only=True)
        val = str(wb[sheet]["A1"].value or "").replace(prefixo, "").strip()
        wb.close()
        return val
    except Exception:
        return caminho.stem


def _titulo(ws, texto: str, n_cols: int, row: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = _TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def _secao(ws, texto: str, n_cols: int, row: int, fill, cor: str = "1F4E79") -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, size=11, color=cor)
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _BORDA
    ws.row_dimensions[row].height = 18


def _cabecalhos(ws, cabecalhos: list[str], row: int) -> None:
    for ci, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=row, column=ci, value=cab)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDA
    ws.row_dimensions[row].height = 30


def _linha(ws, ri: int, row_dict: dict, colunas: list[str],
           colunas_valor: set, colunas_data: set,
           fill, e_novo: bool = False) -> None:
    for ci, col in enumerate(colunas, 1):
        val = row_dict.get(col)
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border = _BORDA
        cell.fill = fill
        cell.font = Font(size=10)
        if col in colunas_data:
            cell.number_format = "dd/mm/yyyy"
            cell.alignment = Alignment(horizontal="center")
        elif col in colunas_valor:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        elif col == "% Repasse":
            cell.number_format = "0.00%"
            cell.alignment = Alignment(horizontal="center")
        elif col == "Novo?":
            cell.alignment = Alignment(horizontal="center")
            if e_novo:
                cell.font = Font(size=10, bold=True, color="7F6000")
        elif col == "Cliente":
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.alignment = Alignment(horizontal="center")


def _subtotal(ws, ri: int, n_cols: int, label: str,
              grupo: pd.DataFrame, colunas: list[str],
              colunas_valor: set, pct: float, pct_ci: int | None,
              e_novo: bool = False) -> None:
    cor = "7F6000" if e_novo else "1F4E79"
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=ri, column=ci)
        c.fill = _SUBTOT_FILL
        c.font = Font(bold=True, size=10, color=cor)
        c.border = _BORDA
    ws.cell(row=ri, column=2, value=f"Subtotal – {label}").alignment = Alignment(horizontal="left")
    for ci, col in enumerate(colunas, 1):
        if col in colunas_valor:
            c = ws.cell(row=ri, column=ci, value=grupo[col].sum())
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
    if pct_ci:
        c = ws.cell(row=ri, column=pct_ci, value=pct)
        c.number_format = "0.00%"
        c.alignment = Alignment(horizontal="center")


def _total(ws, ri: int, n_cols: int, label: str,
           df_sub: pd.DataFrame, colunas: list[str],
           colunas_valor: set, pct: float, pct_ci: int | None,
           fill, cor_fonte: str = "FFFFFF") -> None:
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=ri, column=ci)
        c.fill = fill
        c.font = Font(bold=True, color=cor_fonte, size=10)
        c.border = _BORDA_MED
    ws.cell(row=ri, column=1, value=label).alignment = Alignment(horizontal="center")
    for ci, col in enumerate(colunas, 1):
        if col in colunas_valor:
            c = ws.cell(row=ri, column=ci, value=df_sub[col].sum())
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
    if pct_ci:
        c = ws.cell(row=ri, column=pct_ci, value=pct)
        c.number_format = "0.00%"
        c.alignment = Alignment(horizontal="center")


# ════════════════════════════════════════════════════════════════════════════
# SINTÉTICO
# ════════════════════════════════════════════════════════════════════════════

def adicionar_fechamento_sintetico(caminho: Path) -> dict:
    mes_label, _ = _mes_anterior()
    nome_cc = _nome_cc_da_planilha(caminho, "Consolidado", "Contas Recebidas – ")

    df_raw = pd.read_excel(caminho, sheet_name="Consolidado", header=2)
    df = df_raw[df_raw["Cliente"].notna()].copy()
    df = df[~df["Cliente"].astype(str).str.startswith("TOTAL")]
    df["Novo?"] = df["Novo?"].astype(str).str.strip()

    cols_num_possivel = ["Amortização", "Juros", "Correção", "Acréscimo",
                         "Vl. Baixa", "Seguro", "Taxa adm", "Desconto", "Líquido"]
    cols_num = [c for c in cols_num_possivel if c in df.columns]
    for col in cols_num:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    df["Vlr. Líq. Repasse"] = pd.to_numeric(df["Vlr. Líq. Repasse"], errors="coerce").fillna(0.0)

    colunas_valor = set(cols_num + ["Vlr. Líq. Repasse"])
    pct = float(df["% Repasse"].iloc[0]) if "% Repasse" in df.columns else 0.0
    df_novos = df[df["Novo?"] == "S"].copy()

    colunas = ["Cliente", "Novo?"] + cols_num + ["% Repasse", "Vlr. Líq. Repasse"]
    cabecalhos = colunas[:]
    n_cols = len(colunas)
    pct_ci = next((i for i, c in enumerate(colunas, 1) if c == "% Repasse"), None)
    colunas_data: set = set()

    wb = load_workbook(caminho)
    if "Fechamento" in wb.sheetnames:
        del wb["Fechamento"]
    ws = wb.create_sheet("Fechamento")

    ri = 1
    _titulo(ws, f"Fechamento {mes_label}  –  {nome_cc}", n_cols, ri)
    ri += 1

    # Seção 1 – Todos os clientes
    _secao(ws, f"▶  REPASSE DO MÊS ANTERIOR ({mes_label})  –  Todos os Clientes", n_cols, ri, _SECT_FILL)
    ri += 1
    _cabecalhos(ws, cabecalhos, ri)
    ri += 1

    for idx, (_, row_data) in enumerate(df.iterrows()):
        e_novo = row_data["Novo?"] == "S"
        fill = _NOVO_FILL if e_novo else (_ALT_FILL if idx % 2 == 0 else PatternFill())
        _linha(ws, ri, row_data.to_dict(), colunas, colunas_valor, colunas_data, fill, e_novo)
        ri += 1

    _total(ws, ri, n_cols, "TOTAL GERAL", df, colunas, colunas_valor, pct, pct_ci, _TOTAL_FILL)
    ri += 2

    # Seção 2 – Novos clientes
    if len(df_novos) > 0:
        _secao(ws, f"★  NOVOS CLIENTES – Somatório Acumulado  (CRI mês anterior)", n_cols, ri, _NOVO_FILL, "7F6000")
        ri += 1
        _cabecalhos(ws, cabecalhos, ri)
        ri += 1

        for _, row_data in df_novos.iterrows():
            _linha(ws, ri, row_data.to_dict(), colunas, colunas_valor, colunas_data, _NOVO_FILL, True)
            ri += 1

        _total(ws, ri, n_cols, "TOTAL NOVOS", df_novos, colunas, colunas_valor, pct, pct_ci, _NOVO_TOT_FILL, "000000")
        ri += 1

    # Larguras
    larguras = {"Cliente": 42, "Novo?": 8, "% Repasse": 12, "Vlr. Líq. Repasse": 18}
    for col in cols_num:
        larguras[col] = 14
    for i, col in enumerate(colunas, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 12)
    ws.freeze_panes = "A4"

    wb.save(caminho)
    logger.info("[SINTÉTICO] Aba 'Fechamento' salva → %s", caminho.name)

    return {
        "nome_cc": nome_cc,
        "mes": mes_label,
        "total_repasse": df["Vlr. Líq. Repasse"].sum(),
        "novos": df_novos[["Cliente", "Vlr. Líq. Repasse"]].copy(),
    }


# ════════════════════════════════════════════════════════════════════════════
# ANALÍTICO
# ════════════════════════════════════════════════════════════════════════════

def adicionar_fechamento_analitico(caminho: Path) -> dict:
    mes_label, mes_periodo = _mes_anterior()
    nome_cc = _nome_cc_da_planilha(caminho, "Analítico", "Contas Recebidas Analítico – ")

    df_raw = pd.read_excel(caminho, sheet_name="Analítico", header=2)
    df = df_raw[df_raw["Cliente"].notna()].copy()
    df = df[~df["Cliente"].astype(str).str.startswith("Subtotal")]
    df = df[~df["Cliente"].astype(str).str.startswith("TOTAL")]
    df["Novo?"] = df["Novo?"].astype(str).str.strip()
    df["Dt. Baixa"] = pd.to_datetime(df["Dt. Baixa"], errors="coerce")

    cols_num_possivel = ["Amortização", "Juros", "Correção", "Acréscimo",
                         "Vl. Baixa", "Seguro", "Taxa Adm", "Desconto", "Líquido"]
    cols_num = [c for c in cols_num_possivel if c in df.columns]
    for col in cols_num:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    df["Vlr. Líq. Repasse"] = pd.to_numeric(df["Vlr. Líq. Repasse"], errors="coerce").fillna(0.0)

    colunas_valor = set(cols_num + ["Vlr. Líq. Repasse"])
    colunas_data = {"Dt. Baixa", "Vencimento"}
    pct = float(df["% Repasse"].iloc[0]) if "% Repasse" in df.columns else 0.0

    df_mes = df[df["Dt. Baixa"].dt.to_period("M") == mes_periodo].copy()
    novos_nomes = df[df["Novo?"] == "S"]["Cliente"].unique()
    df_novos_full = df[df["Cliente"].isin(novos_nomes)].copy()

    cols_fixas = [c for c in ["Dt. Baixa", "Cliente", "Novo?", "Documento", "Título",
                              "Parcela", "TC", "Unidade", "Portador", "Operação", "Vencimento"]
                  if c in df.columns]
    colunas = cols_fixas + cols_num + ["% Repasse", "Vlr. Líq. Repasse"]
    cabecalhos = colunas[:]
    n_cols = len(colunas)
    pct_ci = next((i for i, c in enumerate(colunas, 1) if c == "% Repasse"), None)

    wb = load_workbook(caminho)
    if "Fechamento" in wb.sheetnames:
        del wb["Fechamento"]
    ws = wb.create_sheet("Fechamento")

    ri = 1
    _titulo(ws, f"Fechamento {mes_label}  –  {nome_cc}", n_cols, ri)
    ri += 1

    # Seção 1 – Mês anterior
    _secao(ws, f"▶  REPASSE DO MÊS ANTERIOR ({mes_label})  –  Todos os Clientes", n_cols, ri, _SECT_FILL)
    ri += 1
    _cabecalhos(ws, cabecalhos, ri)
    ri += 1

    if len(df_mes) > 0:
        for cliente, grupo in df_mes.groupby("Cliente", sort=True):
            e_novo = (grupo["Novo?"] == "S").any()
            for idx, (_, row_data) in enumerate(grupo.iterrows()):
                fill = _NOVO_FILL if e_novo else (_ALT_FILL if idx % 2 == 0 else PatternFill())
                _linha(ws, ri, row_data.to_dict(), colunas, colunas_valor, colunas_data, fill, e_novo)
                ri += 1
            _subtotal(ws, ri, n_cols, cliente, grupo, colunas, colunas_valor, pct, pct_ci, e_novo)
            ri += 1
        _total(ws, ri, n_cols, "TOTAL GERAL – MÊS ANTERIOR", df_mes, colunas, colunas_valor, pct, pct_ci, _TOTAL_FILL)
        ri += 2
    else:
        ws.merge_cells(f"A{ri}:{get_column_letter(n_cols)}{ri}")
        ws.cell(row=ri, column=1, value="Nenhum registro encontrado para o mês anterior.").font = Font(italic=True,
                                                                                                       color="595959")
        ri += 2

    # Seção 2 – Novos clientes série histórica
    if len(df_novos_full) > 0:
        _secao(ws, f"★  NOVOS CLIENTES – Série Histórica Completa  (CRI mês anterior)", n_cols, ri, _NOVO_FILL,
               "7F6000")
        ri += 1
        _cabecalhos(ws, cabecalhos, ri)
        ri += 1

        for cliente, grupo in df_novos_full.groupby("Cliente", sort=True):
            for _, row_data in grupo.sort_values("Dt. Baixa").iterrows():
                _linha(ws, ri, row_data.to_dict(), colunas, colunas_valor, colunas_data, _NOVO_FILL, True)
                ri += 1
            _subtotal(ws, ri, n_cols, cliente, grupo, colunas, colunas_valor, pct, pct_ci, True)
            ri += 1

        _total(ws, ri, n_cols, "TOTAL NOVOS CLIENTES", df_novos_full, colunas, colunas_valor, pct, pct_ci,
               _NOVO_TOT_FILL, "000000")
        ri += 1

    # Larguras
    larguras = {
        "Dt. Baixa": 12, "Cliente": 40, "Novo?": 7, "Documento": 16,
        "Título": 14, "Parcela": 10, "TC": 8, "Unidade": 14,
        "Portador": 12, "Operação": 12, "Vencimento": 12,
        "% Repasse": 12, "Vlr. Líq. Repasse": 18,
    }
    for col in cols_num:
        larguras[col] = 14
    for i, col in enumerate(colunas, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 12)
    ws.freeze_panes = "A4"

    wb.save(caminho)
    logger.info("[ANALÍTICO] Aba 'Fechamento' salva → %s", caminho.name)

    novos_subtot = (
        df_novos_full.groupby("Cliente")["Vlr. Líq. Repasse"].sum().reset_index()
        if len(df_novos_full) > 0 else pd.DataFrame()
    )
    return {
        "nome_cc": nome_cc,
        "mes": mes_label,
        "total_mes": df_mes["Vlr. Líq. Repasse"].sum() if len(df_mes) > 0 else 0.0,
        "novos": novos_subtot,
    }


# ════════════════════════════════════════════════════════════════════════════
# ENTRYPOINT
# ════════════════════════════════════════════════════════════════════════════

def _encontrar_arquivos(padrao: str) -> list[Path]:
    base = Path("./files/output/contas_recebidas/")
    return list(base.rglob(padrao))


def main() -> None:
    sinteticos = _encontrar_arquivos("*_consolidado_sintetico.xlsx")
    analiticos = _encontrar_arquivos("*_consolidado_analitico.xlsx")

    resultados: list[dict] = []

    # ── Sintéticos ─────────────────────────────────────────────
    for path_sin in sinteticos:

        try:
            logger.info("Processando sintético: %s", path_sin)
            res = adicionar_fechamento_sintetico(path_sin)

            resultados.append({
                "tipo": "sintético",
                **res
            })

        except Exception:
            logger.exception("Erro no sintético: %s", path_sin)

    # ── Analíticos ─────────────────────────────────────────────
    for path_ana in analiticos:

        try:
            logger.info("Processando analítico: %s", path_ana)

            res = adicionar_fechamento_analitico(path_ana)

            resultados.append({
                "tipo": "analítico",
                **res
            })

        except Exception:
            logger.exception("Erro no analítico: %s", path_ana)

    if not resultados:
        logger.error("Nenhum arquivo processado.")
        return

    # ── Resumo WhatsApp consolidado ───────────────────────────
    print("\n" + "=" * 70)
    print("RESUMO")
    print("=" * 70)

    linhas = []

    for r in resultados:

        nome_cc = r["nome_cc"]
        mes = r["mes"]

        if r["tipo"] == "analítico":
            linhas.append(f"*{nome_cc}*")
            linhas.append(f"Competência: {mes}")

            total = r.get("total_mes", 0)

            linhas.append(
                f"Repasse do mês fechado: *R$ {total:,.2f}*"
            )

        # else:
        #
        #     total = r.get("total_repasse", 0)
        #
        #     linhas.append(
        #         f"Total acumulado: *R$ {total:,.2f}*"
        #     )

        # novos = r.get("novos")
        #
        # if novos is not None and len(novos) > 0:
        #
        #     linhas.append("")
        #     linhas.append("*Novos clientes:*")
        #
        #     col_vlr = (
        #         "Vlr. Líq. Repasse"
        #         if "Vlr. Líq. Repasse" in novos.columns
        #         else novos.columns[-1]
        #     )
        #
        #     col_cli = (
        #         "Cliente"
        #         if "Cliente" in novos.columns
        #         else novos.columns[0]
        #     )
        #
        #     for _, row in novos.iterrows():
        #
        #         cliente = str(row[col_cli]).split("(")[0].strip()
        #         valor = row[col_vlr]
        #
        #         linhas.append(
        #             f"• {cliente}: R$ {valor:,.2f}"
        #         )

        linhas.append("")
        linhas.append("-" * 40)
        linhas.append("")

    print("\n".join(linhas))
    print("=" * 70)


if __name__ == "__main__":
    main()
